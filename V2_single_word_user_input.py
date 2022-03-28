#For extracting all lines where a specific word occurs, writing it to a csv in the form [transcription_file_name, row_in_file, original_text]

import os
import numpy as np
import nltk
import re
import spacy
from spacy.symbols import ORTH
import matplotlib
import matplotlib.pyplot as plt
import pandas
import sys
import csv
from openpyxl import load_workbook

SPACY_MODEL = spacy.load("en_core_web_lg")
WORD_OF_INTEREST = str(input("please type a single word you would like to look at: "))
CSV_NAME = str(input("please name your csv file (including the .csv extension!): "))
WORD_DICT = {}
MY_STOPWORDS = ["?","!" ,"@" ,"," ,"-" ,"(", ")" ,"[", "]", ";", ":", "?", ",",'']
special_case = [{ORTH: "like*"}]
SPACY_MODEL.tokenizer.add_special_case("like*", special_case)

def get_file_names(): #names of all files that are English transcripts to be used later
    files_list = []
    current_wd = os.getcwd()
    os.chdir(current_wd+"/workbooks")
    files = os.listdir()
    for file in files:
        if file[0] != '.' and file[0] == 'E': #don't want the .DS store thing or any of the output csvs if you've already run this
            files_list.append(file)
    return files_list

def create_TextFile_classes(files_list): #creates class objects of each file & puts them in a list to be used for easy access later
    file_class_list = []
    for file in files_list:
        file_name = file
        file_class = TextFile(file_name)
        file_class_list.append(file_class)
    return file_class_list

def clean_up(text_file):
    text = text_file.strip("\n")
    text = text.replace('\\','') #could't make this work with regular expressions, got a bad escape error
    text = re.sub("[\{].*?[\}]", " ", text) #removing everything in brackets, parenthese, or curly brackets
    text = re.sub("[\[].*?[\]]","", text)
    text = re.sub("[\(].*?[\)]","", text)
    text = re.sub("[\{].*?[\}]","",text)
    text = re.sub("[［].*?[］]","", text)
    text = re.sub("[\[].*?[\]]","", text)
    text = re.sub("[【].*?[】]","", text)
    text = re.sub("[\[].*?[\]]","",text)
    text = re.sub("[【].*? [】]]","",text)
    text = re.sub("[\[\［\[\【\「\[].*?[\]\】\]\］\」\］]","",text)
    text = re.sub("[｛].*?[］]","",text)
    text = re.sub("[\(\（].*?[\）\）]","",text)
    text = re.sub("[\｛].*?[\｝]","",text)
    text = re.sub("[\*]{3,}","", text) #removing 3-4 sequential stars
    text = re.sub("[\#]","", text) #this … shows up sometimes, is different from ...
    text = re.sub("'","'",text) #??? what is that first character doing
    #text = re.sub("\."," ", text)
    text = re.sub("。","",text)
    #text = re.sub("…","_",text)
    text = re.sub("sss","",text)
    text = re.sub("xpp","",text)
    text = re.sub("xpx","",text)
    text = re.sub("xmm","",text)
    text = re.sub("jjj","",text)
    text = re.sub("jxj","",text)
    text = re.sub("fxf","",text)
    text = re.sub("mmm","",text)
    text = re.sub("mxm","",text)
    text = re.sub("bbb","",text)
    text = re.sub("vvv","",text)
    text = re.sub("ccc","",text)
    text = re.sub("xkx","",text)
    text = re.sub("kxk","",text)
    text = re.sub("xaa","",text)
    text = re.sub("xax","",text)
    text = re.sub("axa","",text)
    text = re.sub("rrr","",text)
    text = re.sub("txt","",text)
    text = re.sub("xtt","",text)
    text = re.sub("ttt","",text)
    text = re.sub("SSS","",text)
    text = re.sub("fff","",text)
    text = re.sub("aaa","",text)
    text = re.sub("eee","",text)
    text = re.sub("xjx","",text)
    text = re.sub("xxss","",text)
    text = re.sub("xss","",text)
    text = re.sub("mmx","",text)
    text = re.sub("xvx","",text)
    text = re.sub("xvv","",text)
    text = re.sub("xff","",text)
    text = re.sub("xsx","",text)
    text = re.sub("sxs","",text)
    text = re.sub("ssx","",text)
    text = re.sub("sss","",text)
    #text = re.sub(":","",text)
    #text = re.sub("!","",text)
    #text = re.sub("/","",text)
    text = re.sub("bxb","",text)
    text = re.sub("\n","", text) #removes newline characters because for some reason that was still an issue even with the first line
    text = re.sub('"*',"",text) #removes multiple quotation characters
    text = re.sub("\s+"," ", text) #removes multiple spaces and replaces with one space
    #text = re.sub("\?","",text)
    text = re.sub("é","",text)
    text = re.sub("õ", "", text)
    #text = re.sub(",","",text)
    #text = re.sub("’","'", text)
    text = re.sub("[xx]{2,}","", text) #removing any number of sequential xs
    text = re.sub(" ss"," ",text)
    text = re.sub(" - "," ",text)
    text = re.sub("\s+"," ", text) #removes multiple spaces and replaces with one space
    tokenized_text = text.split(" ")
    final = [word.lower() for word in tokenized_text if word not in MY_STOPWORDS]
    return final #clean up function. Will vary based on purpose of code (see code file name)

class TextFile:
    def __init__(self, file_name):
        self.file_name = file_name
        self.word_dict = {}

    def read_file(self,file_name): #reads the excel files - file name comes from files_list. Also has the dictionaries for the specific words you want. You'll need to change this if you want to look at more/different words
        file_name = file_name
        lines_list = []
        j = 0

        wb = load_workbook(filename = file_name, read_only=True) #load the excel file using openxyl
        ws = wb['Sheet1'] #go to sheet1 specifically
        for i in ws.iter_rows(min_row=2, min_col=5, max_col=5): #start at second row to skip headers, go down column 5 for transcript
            j += 1
            for cell in i:
                E_text_cell = cell.value
                if cell.value != None and len(E_text_cell) > 0: #if there's words in the cell:
                    E_clean_text = clean_up(E_text_cell) #cleans up the raw text, returns a list where each element is a word
                    clean_text_string = " ".join(E_clean_text)
                    if len(E_clean_text) > 0: #some are empty after cleaning, don't need those
                        E_clean_tokenized_text = SPACY_MODEL(clean_text_string) #turns the list to a string, tags it using spacy
                        excel_cell_as_list_cleaned = [token.text for token in E_clean_tokenized_text]
                        lines_list.append(excel_cell_as_list_cleaned)
                        if WORD_OF_INTEREST in excel_cell_as_list_cleaned:
                            self.word_dict[j] = E_text_cell #dictionary of the line number and the text associated with it for the participant
        print(self.word_dict)
        return self.word_dict

def change_or_make_path(path_addition):
    if os.path.exists(os.getcwd()+"/"+path_addition):
        os.chdir(os.getcwd()+"/"+path_addition)
    else:
        os.mkdir(path_addition)
        os.chdir(os.getcwd()+"/"+path_addition)

def nested_dictionary_csv_writer(dictionary, file_name, label_list):
    #key is the file name
    #value is the dictionary with the lines the target word appeared in
    with open(file_name,"w",newline="") as f:
        thewriter = csv.writer(f)
        thewriter.writerow(label_list)
        for item in dictionary.items():
            file_name = item[0]
            inner_dictionary = item[1]
            for kvp in inner_dictionary.items():
                row_number = kvp[0]
                words = kvp[1]
                thewriter.writerow([file_name,row_number,words])#If your data structure is a nested csv, use this one instead

def main():
    files_list = get_file_names() #gets the excel file names to read
    files_class_list = create_TextFile_classes(files_list) #creating class instances for the text files
    for file_class in files_class_list: #for each class instance
        name = file_class.file_name #file name so it knows which file to read
        word_lines = file_class.read_file(name)
        WORD_DICT[name] = word_lines
    change_or_make_path("/specific_words_individual")
    nested_dictionary_csv_writer(WORD_DICT, CSV_NAME, ["file_name","utterance_number","text"])

main()
