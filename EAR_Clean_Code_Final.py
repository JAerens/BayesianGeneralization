#Project program script
import os
import numpy as np
import re
import matplotlib
import matplotlib.pyplot as plt
import pandas
import seaborn as sns
import csv
from openpyxl import load_workbook #prior versions used xlrd, which at some point decided to only read xls files and no longer support xlsx. That is our data file format, hence the change

PUNCT = ["?","!" ,"@" ,",","(", ")" ,"[", "]", ";", ":", "?", ",",'']
#N_WORDS = int(input("Please select the number of words you'd like to use for things like graphs. Recommended = 20. Number : "))
#WINDOW_SIZE = int(input("Please select the window size you would like to use for measuring contextual diversity: "))

LABEL_LIST = ["Participant","Total_English_Types","Total_Non-English_Types", "English_Tokens", "Non-English_Tokens", "Total_Tokens", "Englis_Types_Per_500_Tokens", "Non-English_Types_Per_500_Tokens", "English_Types_Per_1000_Tokens", "Non-English_Types_Per_1000_Tokens", "Englis_Types_Per_2000_Tokens", "Non-English_Types_Per_2000_Tokens", "Number_Of_English_Speech_Files", "Number_Of_Non-English_Speech_Files", "Average_Number_Of_Words_Per_English_Speech_Sample","Average_Number_Of_Words_Per_Non-English_Speech_Sample", "Number_Of_Valid_Speech_Files","Double_checking_number_of_valid_speech_files", "Number_Of_Files_With_Speech_Total_Not_Double-counted" ,"Total_Number_Of_Weekday_Files_Recorded","Number_of_Valid_Weekday_Files","Number_Of_Total_Weekend_Files_Recorded","Total_Valid_Weekend_Files","E_Tokens_Weekday", "E_Types_Weekday", "NE_Tokens_Weekday", "NE_Types_Weekday", "E_Tokens_Weekend", "E_Types_Weekend", "NE_Tokens_Weekend", "NE_Types_Weekend", "E_NE_Tokens_Weekday", "E_NE_Types_Weekday", "E_NE_Tokens_Weekend", "E_NE_Types_Weekend", "Weekday_English_Files_with_actual_speech", "Weekend_English_Files_with_actual_speech", "Weekday_actual_speech_files_with_only_English_speech", "Weekend_actual_speech_files_with_only_English_speech", "Number_of_weekday_files_with_any_non-english_speech", "Number_of_weekend_files_with_any_non-english_speech", "Number_of_weekday_files_with_exclusively_non-english_speech", "Number_of_weekend_files_with_exclusively_non-english_speech", "Number_of_weekday_files_with_both_english_and_non-english", "Number_of_weekend_files_with_both_english_and_non-english", "Overall_week_speech_files_either_lang","Overall_weekend_speech_files_either_lang", "ToMOnlyFiles","ToFOnlyFiles","ToBothMFiles", "ToSelfFiles", "ToKnownFiles", "ToStrangerFiles", "ToChildFiles", "ToPetFiles", "FilesWithNoConvoPartnerInfo", "OutdoorFiles", "AptDormFiles", "ClassroomFiles", "InTransitVehicleFiles","InTransitOtherFiles","BarCoffeeShopRestaurantFiles","ShoppingFiles","OtherPublicPlacesFiles","NoLocationInDataFiles"]

#Stats to calculate per participant
E_Tokens_Dict = {} #Tokens in English speech
E_Types_Dict = {} #Types in English speech
NE_Tokens_Dict = {} #Tokens in Non-English speech
NE_Types_Dict = {} #Types in Non-English speech
E_NE_Tokens_Dict = {} #Combined tokens
E_NE_Types_Dict = {} #Combined types
E_NE_NoDouble_Speech_Dict = {} #number of files that had speech (either language, not double counting codeswitching)
Valid = {} #Number of valid speech files
Valid_check = {} #just checking somethign

E_500_Dict = {} #Types contained in the first 500 tokens
E_1000_Dict = {} #Types contained in the first 1,000 tokens of English speech
E_2000_Dict = {} #Tokens contained in the first 1,000 tokens of English speech
E_Num_Files_With_Speech_Dict = {} #English number of files containing speech
E_Avg_Num_Words_Per_File_Dict = {} #English average number of words per file

NE_500_Dict = {}
NE_1000_Dict = {}
NE_2000_Dict = {}
NE_Num_Files_With_Speech_Dict = {}
NE_Avg_Num_Words_Per_File_Dict = {}

#weekday vs weekend measures
E_Weekday_Tokens_Dict = {}
E_Weekday_Types_Dict = {}
NE_Weekday_Tokens_Dict = {}
NE_Weekday_Types_Dict = {}
E_Weekend_Tokens_Dict = {}
E_Weekend_Types_Dict = {}
NE_Weekend_Tokens_Dict = {}
NE_Weekend_Types_Dict = {}

#Totals across E & NE for each participant for weekend & weekday
E_NE_Weekday_Tokens_Dict = {}
E_NE_Weekday_Types_Dict = {}
E_NE_Weekend_Tokens_Dict = {}
E_NE_Weekend_Types_Dict = {}

#Total files per participant, speech or no speech
Total_Week_Files_Dict = {}
Total_Weekend_Files_Dict = {}

#Total VALID FILES per participant, speech or no speech
Total_Valid_Week_Files_Dict = {}
Total_Valid_Weekend_Files_Dict = {}

#Overall files across weekdays or weekends that contain speech
Overall_Total_Week_Speech_Files = {}
Overall_Total_Weekend_Speech_Files = {}

#Total number of weekend and weekday files in each E or NE - just number of valid speech files in each case
Total_E_Week_Files_With_Speech = {}
Total_E_Weekend_Files_With_Speech = {}
Total_NE_Week_Files_With_Speech = {}
Total_NE_Weekend_Files_With_Speech = {}

#Number of files with speech where participants codeswitched
Weekday_Speech_Files_Both = {}
Weekend_Speech_Files_Both = {}

#Number of files with only English or only Non-English Speech
Only_E_Week_Files_With_Speech = {}
Only_E_Weekend_Files_With_Speech = {}
Only_NE_Week_Files_With_Speech = {}
Only_NE_Weekend_Files_With_Speech = {}

#Conversation Partner information added 3/14/2022
ToM_Dict = {}
ToF_Dict = {}
ToMF_Dict = {}
ToSelf_Dict = {}
ToKnownPerson_Dict = {}
ToStranger_Dict = {}
ToChild_Dict = {}
ToPet_Dict = {}
NoInfoConversationPartner_Dict = {}

#Location information added 3/14
OutdoorFiles_Dict = {}
AptDorm_Dict = {}
Classroom_Dict = {}
InTransitVehicle_Dict = {}
InTransitOther_Dict = {}
BarCoffeeShopRestaurant_Dict = {}
Shopping_Dict = {}
OtherPublicPlaces_Dict = {}
NoLocation_Dict ={}

def get_file_names():
    '''gets the name of all your text/excel files to be read in later'''
    files_list = []
    os.chdir(os.getcwd()+"/workbooks")
    files = os.listdir()
    for file in files:
        if file[0] != '.' and file[0] == 'E': #checks that the files are not .DS_Store and are all some form of EAR_PARTICIPANT#
            files_list.append(file) #if so, add it
    return files_list #list of all the participant excel files (each participant has their own excel file)

def create_TextFile_classes(files_list):
    file_class_list = []
    for file in files_list:
        file_name = file
        file_class = TextFile(file_name) #make the class instance
        file_class_list.append(file_class) #add it to your list
    return file_class_list #list of the participant files as class instances

def clean_up(text_file):
    '''Function to clean text input. I'm sorry it's a mess. Change lines as needed based on your analyses and coding scheme.
    Also yes I know there's a bunch of regular expressions that can be combined but for our purposes, L2 speech was coded as a 3 character string of a bunch of different consonants, so removing any combination of the letters in the tags led to segments of words being deleted :('''

    text = text_file.replace("\n"," ")
    text = text.replace('\\','') #could't make this work with regular expressions, got a bad escape error
    text = re.sub("[\[].*?[\]]", "", text) #removing everything in brackets, parenthese, or curly brackets
    text = re.sub("[\{].*?[\}]","", text)
    text = re.sub("[\(].*?[\)]","", text)
    text = re.sub("[\<].*?[\>]","", text)
    text = re.sub("[\[\{].*?[\}\]]","",text)
    text = re.sub("sss","",text)
    text = re.sub("sxx","",text)
    text = re.sub("tss","",text)
    text = re.sub("rrr","",text)
    text = re.sub("kkk","",text)
    text = re.sub("ppp","",text)
    text = re.sub("eee","",text)
    text = re.sub("cxc","",text)
    text = re.sub("ggg","",text)
    text = re.sub("mmm","",text)
    text = re.sub("Sxs","",text)
    text = re.sub("Sss","",text)
    text = re.sub("Ccc","",text)
    text = re.sub("xpp","",text)
    text = re.sub("xpx","",text)
    text = re.sub("xrr","",text)
    text = re.sub("xmm","",text)
    text = re.sub("jjj","",text)
    text = re.sub("jxj","",text)
    text = re.sub("fxf","",text)
    text = re.sub("mmm","",text)
    text = re.sub("mxm","",text)
    text = re.sub("bbb","",text)
    text = re.sub("vvv","",text)
    text = re.sub("ccc","",text)
    text = re.sub("xkx","",text)
    text = re.sub("kxk","",text)
    text = re.sub("xaa","",text)
    text = re.sub("xax","",text)
    text = re.sub("axa","",text)
    text = re.sub("rrr","",text)
    text = re.sub("txt","",text)
    text = re.sub("xtt","",text)
    text = re.sub("ttt","",text)
    text = re.sub("SSS","",text)
    text = re.sub("fff","",text)
    text = re.sub("aaa","",text)
    text = re.sub("eee","",text)
    text = re.sub("xjx","",text)
    text = re.sub("xxss","",text)
    text = re.sub("xss","",text)
    text = re.sub("mmx","",text)
    text = re.sub("xvx","",text)
    text = re.sub("xvv","",text)
    text = re.sub("xff","",text)
    text = re.sub("xsx","",text)
    text = re.sub("sxs","",text)
    text = re.sub("ssx","",text)
    text = re.sub("sss","",text)
    text = re.sub("Ttt","",text)
    text = re.sub("Aaa","",text)
    text = re.sub("Eee","",text)
    text = re.sub("bxb","",text)
    text = re.sub("6","six",text)
    text = re.sub("3","three",text)
    text = re.sub(" 2 "," two ",text)
    text = re.sub("4","four",text)
    text = re.sub("5","five",text)
    text = re.sub(" 1 "," one ",text)
    text = re.sub("21st","twenty-first",text)
    text = re.sub(" '", " ", text)
    text = re.sub("-"," ",text)
    text = re.sub(";","",text)
    text = re.sub("[\*]{3,}","", text) #removing all stars (change to {3,} if you want to keep the like*)
    text = re.sub("[\#X]","", text)
    text = re.sub("'","'",text) #??? what is that first character doing
    text = re.sub("\."," ", text)
    text = re.sub("。","",text)
    text = re.sub("…"," ",text)
    text = re.sub("' ", " ", text)
    text = re.sub(":","",text)
    text = re.sub("!","",text)
    text = re.sub("/","",text)
    text = re.sub("\n","", text) #removes newline characters because for some reason that was still an issue even with the first line
    text = re.sub("\?","",text)
    text = re.sub("é","e",text) #some people transcribed words like resume with an accent and excel did not like that
    text = re.sub(",","",text)
    text = re.sub("’","'", text)
    text = re.sub('"',"", text)
    #text = re.sub(r"[sfxvmjrkbctgpSFXVMJRKBCTGP]{3,}", " ", text)
    text = re.sub("[xx]{2,}"," ", text) #removing any number of sequential xs, as those were how conversation partner speech was noted
    text = re.sub("[aAxX]{3,}","",text)
    text = re.sub("[Aa]{3,}","",text)
    text = re.sub("[Ee]{3,}","",text)
    text = re.sub("'", " '", text) #hacky way of separating words with 's
    text = re.sub("\s+"," ", text) #removes multiple spaces and replaces with one space
    tokenized_text = text.split(" ")
    final = [word.lower().strip(" ") for word in tokenized_text if word not in PUNCT]
    return final

def clean_up_NE(text_file):
    ''' text cleaning for Non-English translations text. Varries slightly from the cleaning function for English transcriptions due to transcription code differences'''
    text = text_file.replace("\n"," ")
    text = text.replace('\\','') #could't make this work with regular expressions, got a bad escape error
    text = re.findall("\[(.+?)\]", text)
    text = ' '.join(text)
    text = re.sub("[\{\(].*?[\}\)]", "", text)
    text = re.sub("[\<].*?[\>]","", text)
    text = re.sub("[\*]{3,}","",text)
    text = re.sub("[\#X]","", text) #this … shows up sometimes, is different from ...
    text = re.sub("'","'",text) #??? what is that first character doing
    text = re.sub("\."," ", text)
    text = re.sub("。","",text)
    text = re.sub("…","",text)
    text = re.sub("sss","",text)
    text = re.sub("xpp","",text)
    text = re.sub("xpx","",text)
    text = re.sub("xmm","",text)
    text = re.sub("jjj","",text)
    text = re.sub("jxj","",text)
    text = re.sub("fxf","",text)
    text = re.sub("mmm","",text)
    text = re.sub("mxm","",text)
    text = re.sub("bbb","",text)
    text = re.sub("vvv","",text)
    text = re.sub("ccc","",text)
    text = re.sub("xkx","",text)
    text = re.sub("kxk","",text)
    text = re.sub("xaa","",text)
    text = re.sub("xax","",text)
    text = re.sub("axa","",text)
    text = re.sub("rrr","",text)
    text = re.sub("txt","",text)
    text = re.sub("xtt","",text)
    text = re.sub("ttt","",text)
    text = re.sub("SSS","",text)
    text = re.sub("fff","",text)
    text = re.sub("aaa","",text)
    text = re.sub("eee","",text)
    text = re.sub("xjx","",text)
    text = re.sub("xxss","",text)
    text = re.sub("xss","",text)
    text = re.sub("mmx","",text)
    text = re.sub("xvx","",text)
    text = re.sub("xvv","",text)
    text = re.sub("xff","",text)
    text = re.sub("xsx","",text)
    text = re.sub("sxs","",text)
    text = re.sub("ssx","",text)
    text = re.sub("sss","",text)
    text = re.sub(":","",text)
    text = re.sub("!","",text)
    text = re.sub("/","",text)
    text = re.sub("bxb","",text)
    text = re.sub("\n","", text) #removes newline characters because for some reason that was still an issue even with the first line
    text = re.sub("\?","",text)
    text = re.sub(",","",text)
    text = re.sub("’","'", text)
    text = re.sub("[xx]{2,}","", text) #removing any number of sequential xs
    text = re.sub(" - "," ",text)
    text = re.sub("\s+"," ", text) #removes multiple spaces and replaces with one space
    tokenized_text = text.split(" ")
    final = [word.lower() for word in tokenized_text if word not in PUNCT]
    return final

def make_list_a_string(list_of_lists):
    '''Converting lists to strings for possible future use with Spacy'''

    text_string = ""
    for individual_list in list_of_lists:
        for word in individual_list:
            text_string = text_string + word + " "
    return text_string#for taking the list and turning it into a string in case you need the string for Spacy or something like that

def stacked_grapher(E_data_dict, NE_data_dict, title, x_lab, y_lab):
    '''for making the graphs that have the English & Non-English types & tokens stacked in one bar for each of the participants'''

    #dictionary data needs to be in the form of participants:counts
    participants = list(E_data_dict.keys()) #your participant IDS
    E_counts = list(E_data_dict.values()) #their counts of English types/tokens
    NE_counts = list(NE_data_dict.values()) #their counts of Non-English types/tokens
    sums = []
    for i in range(len(E_counts)):
        sums.append(E_counts[i]+NE_counts[i]) #total of both sources of types/tokens (height of the bar)
    df = pandas.DataFrame() #make an empty dataframe to start
    df['participants'] = participants #participants column
    df['English'] = E_counts #column of English counts alone
    df['Non-English'] = NE_counts #column of Non-English counts alone
    df['Total'] = sums #column of them summed together
    df = df.set_index('participants') #do this otherwise it assigns an index starting at 0 for participants, and if that doesn't correspond to your labeling scheme your collaborators will get confused
    df.sort_values('Total', ascending=True)[['English','Non-English']].plot.bar(stacked=True) #sort it by lowest to highest b/c it's easier to look at than 76 random bars
    plt.title(title, fontsize=36, pad = 20) #formatty stuff
    plt.xlabel('')
    plt.legend(loc="upper left", fontsize = 24)
    plt.ylabel(y_lab, fontsize=24, labelpad = 20)
    plt.tick_params(axis="y", labelsize = 20)
    plt.xticks(rotation=90)
    nme = title+".png"
    plt.savefig(nme)
    plt.show()

def stacked_grapher_general(data_dict_1, data_dict_2, title, x_lab, y_lab, dd1_name, dd2_name):
    '''for making the graphs that have the English & Non-English types & tokens stacked in one bar for each of the participants'''

    #dictionary data needs to be in the form of participants:counts
    participants = list(data_dict_1.keys()) #your participant IDS
    variable1_counts = list(data_dict_1.values()) #their counts of English types/tokens
    variable2_counts = list(data_dict_2.values()) #their counts of Non-English types/tokens
    sums = []
    for i in range(len(variable1_counts)):
        sums.append(variable1_counts[i]+variable2_counts[i]) #total of both sources of types/tokens (height of the bar)
    df = pandas.DataFrame() #make an empty dataframe to start
    df['participants'] = participants #participants column
    df[dd1_name] = variable1_counts #column of English counts alone
    df[dd2_name] = variable2_counts #column of Non-English counts alone
    df['Total'] = sums #column of them summed together
    df = df.set_index('participants') #do this otherwise it assigns an index starting at 0 for participants, and if that doesn't correspond to your labeling scheme your collaborators will get confused
    df.sort_values('Total', ascending=True)[[dd1_name,dd2_name]].plot.bar(stacked=True) #sort it by lowest to highest b/c it's easier to look at than 76 random bars
    plt.title(title, fontsize=36, pad = 20) #formatty stuff
    plt.xlabel('')
    plt.legend(loc="upper left", fontsize = 24)
    plt.ylabel(y_lab, fontsize=24, labelpad = 20)
    plt.tick_params(axis="y", labelsize = 20)
    plt.xticks(rotation=90)
    nme = title+".png"
    plt.savefig(nme)
    plt.show()

def stacked_grapher_proportion(E_week_dict, E_weekend_dict, NE_week_dict, NE_weekend_dict, num_bars, title, y_lab, figname):
    '''for making the graphs that have the English & Non-English types & tokens stacked in one bar for each of the participants'''
    palette = sns.color_palette("colorblind", num_bars)
    #dictionary data needs to be in the form of participants:[counts,total_speech]
    participant = list(E_week_dict.keys())
    Eweek = [i for i in list(E_week_dict.values())]
    Eweekend = [i for i in list(E_weekend_dict.values())]
    NEweek = [i for i in list(NE_week_dict.values())]
    NEweekend = [i for i in list(NE_weekend_dict.values())]
    sums = [Eweek[i]+Eweekend[i]+NEweek[i]+NEweekend[i] for i in range(len(participant))]

    df = pandas.DataFrame() #make an empty dataframe to start
    df['participant'] = participant #participants column
    df['English Week'] = Eweek #column of English counts alone
    df['English Weekend'] = Eweekend #column of Non-English counts alone
    df['NonEnglish Week'] = NEweek #column of Non-English counts alone
    df['NonEnglish Weekend'] = NEweekend #column of Non-English counts alone
    df['Total'] = sums #column of them summed together
    df = df.set_index('participant') #do this otherwise it assigns an index starting at 0 for participants, and if that doesn't correspond to your labeling scheme your collaborators will get confused
    df.sort_values('Total', ascending=True)[['English Week','English Weekend','NonEnglish Week', 'NonEnglish Weekend']].plot.bar(stacked=True, color=palette) #sort it by lowest to highest b/c it's easier to look at than 76 random bars
    plt.title(title, fontsize=36, pad = 20) #formatty stuff
    plt.xlabel('')
    plt.legend(loc="upper left", fontsize = 24)
    plt.ylabel(y_lab, fontsize=24, labelpad = 20)
    plt.tick_params(axis="y", labelsize = 20)
    plt.xticks(rotation=90)
    plt.savefig(figname)
    plt.show()

def stats_csv_writer(name, label_list, Etypesdict, NEtypesdict, Etokensdict, NEtokensdict, bothtokensdict, E_fivehundred, NE_fivehundred, Ethousand_dict, NEthousand_dict, Etwothousand_dict, NEtwothousand_dict, E_Num_FilesWithSpeech_Dict, NE_Num_FilesWithSpeech_Dict, E_Avg_Num_WordsPerFile_Dict, NE_Avg_Num_WordsPerFile_Dict, Valid_Dict, Valid_Dict_Two, SpeechfilesDict, WeekdayfilesDict, WeekdayValidFilesDict, WeekendfilesDict, WeekendValidFilesDict, EWeekday_Tokens_Dict, EWeekday_Types_Dict, NEWeekday_Tokens_Dict, NEWeekday_Types_Dict, EWeekend_Tokens_Dict, EWeekend_Types_Dict, NEWeekend_Tokens_Dict, NEWeekend_Types_Dict, ENE_Weekday_Tokens_Dict, ENE_Weekday_Types_Dict, ENE_Weekend_Tokens_Dict, ENE_Weekend_Types_Dict, E_Week_Files_With_Speech_Dict, E_Weekend_Files_With_Speech_Dict, Only_E_Week_Files_With_Speech_Dict, Only_E_Weekend_Files_With_Speech_Dict, NE_Week_Files_With_Speech_Dict, NE_Weekend_Files_With_Speech_Dict, Only_NE_Week_Files_With_Speech_Dict, Only_NE_Weekend_Files_With_Speech_Dict, BothWeekSpeechFilesDict, BothWeekendSpeechFilesDict, OverallWeekSpeechFiles, OverallWeekendSpeechFiles, ToM, ToF, ToMF, ToSelf, ToKnown, ToStranger, ToChild, ToPet, UnknownToWhom, Outdoors, AptDorm, Classroom, InTransitVehicle, InTransitOther, BarCoffeeShopRestaurant, Shopping, OtherPublicPlaces, NoLocation): #Don't look at it, it's hideous
    '''specific to the data we wanted to look at in particular, which is types, tokens, number of files with speech, average words per speech file, and then types in each of the 3 token set sizes'''

    participants = []
    Etypes = []
    NEtypes = []
    Etokens = []
    NEtokens = []
    alltokens = []
    E_five_hundred = []
    Ethousand = []
    Etwo_thousand = []
    Eavg_words = []
    NE_five_hundred = []
    NEthousand = []
    NEtwo_thousand = []
    Valid_Files = []
    NEavg_words = []
    E_speech_samples = []
    NE_speech_samples = []
    valid_files = []
    valid_files_two = []
    No_Double_Speech = []
    Weekday_Files = []
    Weekday_Valid_Files = []
    Weekend_Files = []
    Weekend_Valid_Files = []
    E_Weekday_Tokens = []
    E_Weekday_Types = []
    NE_Weekday_Tokens = []
    NE_Weekday_Types = []
    E_Weekend_Tokens = []
    E_Weekend_Types = []
    NE_Weekend_Tokens = []
    NE_Weekend_Types = []
    E_NE_Weekday_Tokens = []
    E_NE_Weekday_Types = []
    E_NE_Weekend_Tokens = []
    E_NE_Weekend_Types = []
    E_Week_Files_With_Speech = []
    E_Weekend_Files_With_Speech = []
    Only_E_Week_Files_With_Speech = []
    Only_E_Weekend_Files_With_Speech = []
    NE_Week_Files_With_Speech = []
    NE_Weekend_Files_With_Speech = []
    Only_NE_Week_Files_With_Speech = []
    Only_NE_Weekend_Files_With_Speech = []
    Both_Week_Speech_Files = []
    Both_Weekend_Speech_Files = []
    Overall_Week_Speech_Files = []
    Overall_Weekend_Speech_Files = []

    #Conversaion Partner added 3/14
    num_ToM_files = []
    num_ToF_files = []
    num_ToMF_files = []
    num_ToSelf_files = []
    num_ToKnown_files = []
    num_ToStranger_files = []
    num_ToChild_files = []
    num_ToPet_files = []
    num_NoInfoConversationPartner_files = []

    #location added 3/14
    num_Outdoor_files = []
    num_AptDorm_files = []
    num_Classroom_files = []
    num_InTransitVehicle_files = []
    num_IntransitOther_files = []
    num_BarCoffeeShopRestaurant_files = []
    num_Shopping_files = []
    num_OtherPublicPlaces_files = []
    num_NoLocation_files = []

    #original analyses
    for item in Etypesdict.items():
        participants.append(int(item[0]))
        Etypes.append(item[1])
    for item in NEtypesdict.items():
        NEtypes.append(item[1])
    for item in Etokensdict.items():
        Etokens.append(item[1])
    for item in NEtokensdict.items():
        NEtokens.append(item[1])
    for item in bothtokensdict.items():
        alltokens.append(item[1])
    for item in E_Num_FilesWithSpeech_Dict.items():
        E_speech_samples.append(item[1])
    for item in NE_Num_FilesWithSpeech_Dict.items():
        NE_speech_samples.append(item[1])
    for item in E_Avg_Num_WordsPerFile_Dict.items():
        Eavg_words.append(item[1])
    for item in NE_Avg_Num_WordsPerFile_Dict.items():
        NEavg_words.append(item[1])
    for item in E_fivehundred.items():
        E_five_hundred.append(item[1])
    for item in NE_fivehundred.items():
        NE_five_hundred.append(item[1])
    for item in Ethousand_dict.items():
        Ethousand.append(item[1])
    for item in NEthousand_dict.items():
        NEthousand.append(item[1])
    for item in Etwothousand_dict.items():
        Etwo_thousand.append(item[1])
    for item in NEtwothousand_dict.items():
        NEtwo_thousand.append(item[1])

    #valid file checking
    for item in Valid_Dict.items():
        valid_files.append(item[1])
    for item in Valid_Dict_Two.items():
        valid_files_two.append(item[1])
    for item in SpeechfilesDict.items():
        No_Double_Speech.append(item[1])

    #weekday/weekend
    for item in WeekdayfilesDict.items():
        Weekday_Files.append(item[1])
    for item in WeekdayValidFilesDict.items():
        Weekday_Valid_Files.append(item[1])
    for item in WeekendfilesDict.items():
        Weekend_Files.append(item[1])
    for item in WeekendValidFilesDict.items():
        Weekend_Valid_Files.append(item[1])
    for item in EWeekday_Tokens_Dict.items():
        E_Weekday_Tokens.append(item[1])
    for item in EWeekday_Types_Dict.items():
        E_Weekday_Types.append(item[1])
    for item in NEWeekday_Tokens_Dict.items():
        NE_Weekday_Tokens.append(item[1])
    for item in NEWeekday_Types_Dict.items():
        NE_Weekday_Types.append(item[1])
    for item in EWeekend_Tokens_Dict.items():
        E_Weekend_Tokens.append(item[1])
    for item in EWeekend_Types_Dict.items():
        E_Weekend_Types.append(item[1])
    for item in NEWeekend_Tokens_Dict.items():
        NE_Weekend_Tokens.append(item[1])
    for item in NEWeekend_Types_Dict.items():
        NE_Weekend_Types.append(item[1])
    for item in ENE_Weekday_Tokens_Dict.items():
        E_NE_Weekday_Tokens.append(item[1])
    for item in ENE_Weekday_Types_Dict.items():
        E_NE_Weekday_Types.append(item[1])
    for item in ENE_Weekend_Tokens_Dict.items():
        E_NE_Weekend_Tokens.append(item[1])
    for item in ENE_Weekend_Types_Dict.items():
        E_NE_Weekend_Types.append(item[1])
    for item in E_Week_Files_With_Speech_Dict.items():
        E_Week_Files_With_Speech.append(item[1])
    for item in E_Weekend_Files_With_Speech_Dict.items():
        E_Weekend_Files_With_Speech.append(item[1])
    for item in Only_E_Week_Files_With_Speech_Dict.items():
        Only_E_Week_Files_With_Speech.append(item[1])
    for item in Only_E_Weekend_Files_With_Speech_Dict.items():
        Only_E_Weekend_Files_With_Speech.append(item[1])
    for item in NE_Week_Files_With_Speech_Dict.items():
        NE_Week_Files_With_Speech.append(item[1])
    for item in NE_Weekend_Files_With_Speech_Dict.items():
        NE_Weekend_Files_With_Speech.append(item[1])
    for item in Only_NE_Week_Files_With_Speech_Dict.items():
        Only_NE_Week_Files_With_Speech.append(item[1])
    for item in Only_NE_Weekend_Files_With_Speech_Dict.items():
        Only_NE_Weekend_Files_With_Speech.append(item[1])
    for item in BothWeekSpeechFilesDict.items():
        Both_Week_Speech_Files.append(item[1])
    for item in BothWeekendSpeechFilesDict.items():
        Both_Weekend_Speech_Files.append(item[1])
    for item in OverallWeekSpeechFiles.items():
        Overall_Week_Speech_Files.append(item[1])
    for item in OverallWeekendSpeechFiles.items():
        Overall_Weekend_Speech_Files.append(item[1])

    #location added 3/14
    for item in ToM.items():
        num_ToM_files.append(item[1])
    for item in ToMF.items():
        num_ToF_files.append(item[1])
    for item in ToF.items():
        num_ToMF_files.append(item[1])
    for item in ToSelf.items():
        num_ToSelf_files.append(item[1])
    for item in ToKnown.items():
        num_ToKnown_files.append(item[1])
    for item in ToStranger.items():
        num_ToStranger_files.append(item[1])
    for item in ToChild.items():
        num_ToChild_files.append(item[1])
    for item in ToPet.items():
        num_ToPet_files.append(item[1])
    for item in UnknownToWhom.items():
        num_NoInfoConversationPartner_files.append(item[1])

    #print(ToMF)
    #print(ToSelf)
    #print(ToKnown)
    #print(ToStranger)
    #print(ToChild)
    #print(ToPet)
    #print(UnknownToWhom)

    #location added 3/14
    for item in Outdoors.items():
        num_Outdoor_files.append(item[1])
    for item in AptDorm.items():
        num_AptDorm_files.append(item[1])
    for item in Classroom.items():
        num_Classroom_files.append(item[1])
    for item in InTransitVehicle.items():
        num_InTransitVehicle_files.append(item[1])
    for item in InTransitOther.items():
        num_IntransitOther_files.append(item[1])
    for item in BarCoffeeShopRestaurant.items():
        num_BarCoffeeShopRestaurant_files.append(item[1])
    for item in Shopping.items():
        num_Shopping_files.append(item[1])
    for item in OtherPublicPlaces.items():
        num_OtherPublicPlaces_files.append(item[1])
    for item in NoLocation.items():
        num_NoLocation_files.append(item[1])

    #print(Outdoors)
    #print(AptDorm)
    #print(Classroom)
    #print(InTransitVehicle)
    #print(InTransitOther)
    #print(BarCoffeeShopRestaurant)
    #print(OtherPublicPlaces)
    #print(NoLocation)

    with open(name, "w", newline = "") as f:
        thewriter = csv.writer(f)
        thewriter.writerow(label_list)
        for i in range(len(participants)): #for each person
            thewriter.writerow([str(participants[i]), str(Etypes[i]), str(NEtypes[i]), str(Etokens[i]), str(NEtokens[i]), alltokens[i], E_five_hundred[i], NE_five_hundred[i], Ethousand[i], NEthousand[i], Etwo_thousand[i], NEtwo_thousand[i], E_speech_samples[i], NE_speech_samples[i], Eavg_words[i], NEavg_words[i], valid_files[i], valid_files_two[i], No_Double_Speech[i], Weekday_Files[i], Weekday_Valid_Files[i], Weekend_Files[i], Weekend_Valid_Files[i], E_Weekday_Tokens[i], E_Weekday_Types[i], NE_Weekday_Tokens[i], NE_Weekday_Types[i], E_Weekend_Tokens[i], E_Weekend_Types[i], NE_Weekend_Tokens[i], NE_Weekend_Types[i], E_NE_Weekday_Tokens[i], E_NE_Weekday_Types[i], E_NE_Weekend_Tokens[i], E_NE_Weekend_Types[i], E_Week_Files_With_Speech[i], E_Weekend_Files_With_Speech[i], Only_E_Week_Files_With_Speech[i], Only_E_Weekend_Files_With_Speech[i], NE_Week_Files_With_Speech[i], NE_Weekend_Files_With_Speech[i], Only_NE_Week_Files_With_Speech[i], Only_NE_Weekend_Files_With_Speech[i], Both_Week_Speech_Files[i], Both_Weekend_Speech_Files[i], Overall_Week_Speech_Files[i],Overall_Weekend_Speech_Files[i],num_ToM_files[i],num_ToF_files[i],num_ToMF_files[i],num_ToSelf_files[i],num_ToKnown_files[i],num_ToStranger_files[i],num_ToChild_files[i],num_ToPet_files[i], num_NoInfoConversationPartner_files[i], num_Outdoor_files[i], num_AptDorm_files[i], num_Classroom_files[i], num_InTransitVehicle_files[i], num_IntransitOther_files[i], num_BarCoffeeShopRestaurant_files[i], num_Shopping_files[i], num_OtherPublicPlaces_files[i], num_NoLocation_files[i]])

def create_corpus_dict(all_text_list):
    unique_word_corpus_size = 0
    total_word_corpus_size = 0
    all_text_dict_unique = {}
    all_text_dict_total = {}
    for word in all_text_list:
        total_word_corpus_size += 1
        if word not in all_text_dict_unique:
            all_text_dict_unique[word] = 1 #just making a dictionary that has all the individual words that appear at all in the corpus (so all tokens over all participants)
            unique_word_corpus_size += 1
        else:
            pass
        if word not in all_text_dict_total:
            all_text_dict_total[word] = 1
        else:
            all_text_dict_total[word] += 1
    print(unique_word_corpus_size, " UNIQUE WORDS")
    print(total_word_corpus_size, " TOTAL WORDS")
    return all_text_dict_unique, all_text_dict_total

def make_dict_from_tokens(tokens_in_list_form):
    d = {}
    for t in tokens_in_list_form:
        if t not in d:
            d[t] = 1
        else:
            d[t] += 1
    return d

def contextual_diversity_window(all_words_dictionary, entire_corpus_text_as_a_list, window_size):
    '''calculating the number of unique words that appear around each word in the corpus in a given window size (you set it as input before the program runs)'''

    cd_dict = {} #a dictionary where each key is a word in the corpus, and each value is the list of words that co-occur in the window with that word
    for word in list(all_words_dictionary.keys()): #a dictionary of all the unique words in the corpus, so the keys are your types
        cd_dict[word] = 0 #initialize all counts of how many words are in the window with the word you're currently looking at at zero

    for word_in_dict in list(all_words_dictionary.keys()): #all the types in the entire corpus
        cd_list = [] #list of the words
        indices = [] #list of the indicies
        for index, word_in_list in enumerate(entire_corpus_text_as_a_list): #takes the entire corpus and turns the words into tuples of (the index, the word)
            if word_in_list == word_in_dict: #as you iterate through the corpus, if the word you're on in the list matches the word you're on in the entire types dictionary
                indices.append(index) #add the index of that word in the entire tokens corpus to the list
        for index in indices: #so for all the indicies in the corpus where the word you're lookat at appears
            window_list = entire_corpus_text_as_a_list[index-window_size: index+window_size] #get the words in the window around the word you're looking at
            for each_word in window_list: #for each of those words
                if each_word not in cd_list: #if they're not in the contextual diversity list
                    cd_list.append(each_word) #add them, giving you the total number of different words that appears in the whatever-sized window around any instance of the word you're looking at

        cd_dict[word_in_dict] = len(cd_list) #should be a dictionary of each word in the corpus and the number of words that co-occur with it

    return cd_dict

def contexutal_diversity_participants(all_words_dictionary, class_instances_list):
    '''calculating contextual diversity of each word by how many of the participants said that word in any of their speech files'''

    participant_occurrences_dict = {}
    for word in list(all_words_dictionary.keys()): #for each word in a dictionary of all the types
        participant_occurrences_dict[word] = 0 #start the number of participants who've said each word at zero
        for participant in class_instances_list: #go through all the participant class instances
            their_text = participant.E_one_list #get their text, in a list form
            if word in their_text: #if they said that word
                participant_occurrences_dict[word] += 1 #increase counter
    return participant_occurrences_dict

def all_words_and_cd_csv_writer(file_name, label_list, count_dict, cd_participants_dict, cd_window_dict):
    '''for writing a csv with the word counts, and measures of contextual diversity'''

    #input for keeping track of each row in order
    word_list = []
    counts_list = []
    cd_window_list = []
    cd_participant_list = []

    #putting them into the lists
    for item in count_dict.items():
        word_list.append(item[0])
        counts_list.append(item[1])
    for word in word_list:
        cd_window_list.append(cd_window_dict[word]) #get the number of other words that word appears with in whatever the window was, add it in order the words go in so they match up
        cd_participant_list.append(cd_participants_dict[word]) #same thing for the number of people who said that word (max should be your number of participants)

    #writing the csv file
    with open(file_name, "w", newline = "") as f:
        thewriter = csv.writer(f)
        thewriter.writerow(label_list)
        for i in range(len(word_list)):
            thewriter.writerow([word_list[i],counts_list[i],cd_participant_list[i],cd_window_list[i]])

def words_grapher(csv_file, sort_by_label, title, y_label):
    df = pandas.read_csv(csv_file) #csv with the word counts
    df = pandas.DataFrame(df) #convert to dataframe
    df = df.sort_values(by=sort_by_label, ascending=False) #sort values from largest to smallest
    df = df.head(N_WORDS) #taking just the n most frequent
    plt.bar(x=df['word'], height=df[sort_by_label],color="palevioletred") #make a bar graph of the word counts
    plt.ylabel(y_label)
    plt.xlabel(' ')
    plt.title(title, fontsize=18)
    plt.xticks(rotation = 90)
    plt.legend('',frameon=False)
    plt.show()

def participant_grapher(csv_file, sort_by_label, title, y_label, col):
    df = pandas.read_csv(csv_file, usecols = [0,col]) #col 0 = participant, col = col is the column you want to graph on
    df = pandas.DataFrame(df)
    sorted = df.sort_values(by=sort_by_label)
    #df = df.sort_values('Wordss_Diff_Week_Weekend') #sort values from largest to smallest
    #plt.bar(x='Participant', height='Wordss_Diff_Week_Weekend', data = df, color="palevioletred")
    sorted.plot(x = "Participant", y=sort_by_label, kind="bar")
    #print(df)
    plt.ylabel(y_label, fontsize = 16)
    plt.xlabel('  ')
    plt.legend('',frameon=False)
    plt.title(title, fontsize = 18)
    plt.xticks(rotation = 90, fontsize = 0)
    plt.show()

def proportions_grapher(csv_file, sum_label, t, y_label, col):
    pals = sns.color_palette("Paired", 4)
    df = pandas.read_csv(csv_file, usecols = col) #col 0 = participant, col = col is the column you want to graph on
    df = pandas.DataFrame(df)
    sum = sum_label
    df.plot(x = 'Participant', kind = 'bar', stacked = True, color=pals, ylim=[0,1.2])
    #print(df)
    plt.ylabel(y_label, fontsize = 16)
    #plt.xlabel('  ')
    plt.legend(loc="upper left")
    plt.title(t, fontsize = 18)
    plt.xticks(rotation = 90, fontsize = 0)
    plt.show()

def change_or_make_path(path_addition):
    if os.path.exists(os.getcwd()+"/"+path_addition):
        os.chdir(os.getcwd()+"/"+path_addition)
    else:
        os.mkdir(path_addition)
        os.chdir(os.getcwd()+"/"+path_addition)

def write_corpus_to_text(corpus_as_a_string, file_name):
    with open(file_name, "w") as f:
        f.write(corpus_as_a_string)

class TextFile:
    def __init__(self, file_name):
        self.file_name = file_name

        #each excel cell as a string
        self.E_raw_text_string = None
        self.NE_raw_text_string = None

        #each excel cell as a list
        self.E_raw_text_list = []
        self.NE_raw_text_list = []

        #number of files with speech
        self.num_files_any_speech_E_NE = 0

        #list keeping track of the number of words per each cell (includes empty cells)
        self.E_words_per_chunk_list = []
        self.NE_words_per_chunk_list = []
        self.Total_words_per_chunk_list = []

        #list keeping track of the number of words per cell only when the participant was talking
        self.E_words_per_chunk_with_speech = []
        self.NE_words_per_chunk_with_speech = []
        self.Both_words_per_chunk_with_speech = []

        #type and token counts
        self.E_num_types = 0
        self.E_num_tokens = 0
        self.NE_num_tokens = 0
        self.NE_num_types = 0
        self.types_dict = {}
        self.tokens_dict = {}

        #all participant speech in one big list
        self.E_one_list = []
        self.NE_one_list = []

        #Number of files with participant speech
        self.num_valid_files = 0
        self.lines_E_text = {}
        self.lines_NE_text = {}
        self.lines_total = {}

        #Weekends vs weekdays
        self.num_weekday_files = 0
        self.num_weekend_files = 0
        self.num_valid_weekday_files = 0
        self.num_valid_weekend_files = 0
        self.Overall_week_speech_files = 0
        self.Overall_weekend_speech_files = 0
        #English specifically
        self.Eweek_files = []
        self.Eweekend_files = []
        self.Enum_week_types = 0
        self.Eweek_types_dict = {}
        self.Enum_week_tokens = 0
        self.Enum_weekend_types = 0
        self.Eweekend_types_dict = {}
        self.Enum_weekend_tokens = 0
        self.Enum_weekend_files = 0
        self.Enum_week_files = 0
        self.Enum_week_files_with_speech = 0
        self.Enum_weekend_files_with_speech = 0
        #and now NE
        self.NEweek_files = []
        self.NEweekend_files = []
        self.NEnum_week_types = 0
        self.NEnum_week_tokens = 0
        self.NEweek_types_dict = {}
        self.NEnum_weekend_types = 0
        self.NEnum_weekend_tokens = 0
        self.NEweekend_types_dict = {}
        self.NEnum_weekend_files = 0
        self.NEnum_week_files = 0
        self.NEnum_week_files_with_speech = 0
        self.NEnum_weekend_files_with_speech = 0

        self.valid_counter_two = 0 # just checking something

        self.Num_week_files_with_speech_both = 0
        self.Num_weekend_files_with_speech_both = 0

        self.weekday_only_english_files = 0
        self.weekend_only_english_files = 0
        self.weekday_only_Nonenglish_files = 0
        self.weekend_only_Nonenglish_files = 0

        self.Only_E_Week_Files_With_Speech = 0
        self.Only_E_Weekend_Files_With_Speech = 0
        self.Only_NE_Week_Files_With_Speech = 0
        self.Only_NE_Weekend_Files_With_Speech = 0

        #addition 3/14 - Conversation Partners
        self.num_ToM_files = 0
        self.num_ToF_files = 0
        self.num_ToMF_files = 0
        self.num_ToSelf_files = 0
        self.num_ToKnown_files = 0
        self.num_ToStranger_files = 0
        self.num_ToChild_files = 0
        self.num_ToPet_files = 0
        self.num_NoKnownConvoPartner_files = 0

        #addition 3/13 - locations
        self.num_Outdoor_files = 0
        self.num_AptDorm_files = 0
        self.num_Classroom_files = 0
        self.num_InTransitVehicle_files = 0
        self.num_IntransitOther_files = 0
        self.num_BarCoffeeShopRestaurant_files = 0
        self.num_Shopping_files = 0
        self.num_OtherPublicPlaces_files = 0
        self.num_NoLocation_files = 0

    def read_file(self,file_name): #file name comes from files_list
        file_name = file_name
        #print(file_name)
        E_lines_list = [] #list of list of words in each cell
        E_one_list = [] #all words just as one long list
        NE_lines_list = [] #same for NE
        NE_one_list = []
        Valid_counter = 0
        #print(file_name)
        #The original version of this code used xlrd, but support for xlsx files was dropped before this project reached publicaton. Additional analyses necessitated modifying the code to use openpyxl instead
        #wb = xlrd.open_workbook(file_name, encoding_override="utf-8")
        #sheet = wb.sheet_by_index(0) #first sheet in the workbook
        #for i in range(sheet.nrows): #for each row
            #if i > 0: #but skipping the first row because it just says transcript
        wb = load_workbook(filename = file_name, read_only=True) #load the excel file using openxyl
        ws = wb['Sheet1'] #go to sheet1 specifically
        cell_counter = 0
        start_cell = 1
        row_numeric = 2#needed due to differences in row/column indexing between xlrd and openpyxl

        #this is just to figure out how many rows there are in each spreadsheet so we know where to stop
        #filtering deleted files happens after this step
        for i in ws.iter_rows(min_row = 2, min_col = 3, max_col = 3):
            diff_counter = 0
            #row for file number - openpyxl includes some blank rows that xlrd did not, so making sure it stops at the end of the dataframe
            for cell in i:
                if cell.value != None:
                    if type(cell.value) == int:
                        if cell.value == start_cell:
                            start_cell+=1
                            #print("added 1, start cell now", start_cell)
                        else:
                            #print(start_cell)
                            diff = cell.value-start_cell
                            #print("the difference was: ", diff)
                            diff_counter += diff
                            total_diff = diff+1
                            start_cell += total_diff
                            #print("added the difference, start cell now", start_cell)
                    else:
                        print("File number value not an int", type(cell.value))
                else:
                    break
                cell_counter += 1
            start_cell -= diff_counter
        print("if using cell counter: ", cell_counter, "files for", file_name)
        print("using ", start_cell-1, "rows for file", file_name)
        for i in ws.iter_rows(min_row=2, max_row = start_cell): #start at second row to skip headers, go down column 5 for transcript
            cell_num = ws.cell(row= row_numeric, column = 3).value
            E_text_cell = ws.cell(row = row_numeric, column=5).value # column numbers are one off because this takes column numbers starting at 1, whereas xlrd used python indexing starting at 0
            NE_text_cell = ws.cell(row = row_numeric, column = 6).value
            Sleeping = ws.cell(row_numeric, 34).value #for counting valid files
            Problem = ws.cell(row_numeric, 9).value #also for counting valid files
            Day_of_week = ws.cell(row_numeric, 11).value #weekdays = 1, weekends = 2

            #added conversation partner information 3/14
            MF = ws.cell(row_numeric, 17).value
            Self = ws.cell(row_numeric, 18).value
            KnownPerson = ws.cell(row_numeric, 19).value
            Stranger = ws.cell(row_numeric, 20).value
            Child = ws.cell(row_numeric, 21).value
            Pet = ws.cell(row_numeric, 22).value

            #added location information 3/14
            AptDorm = ws.cell(row_numeric, 38).value
            Classroom = ws.cell(row_numeric, 39).value
            Outdoors = ws.cell(row_numeric, 40).value
            InTransitVehicle = ws.cell(row_numeric, 41).value
            InTransitOther = ws.cell(row_numeric, 42).value
            BarCoffeeShopRestaurant = ws.cell(row_numeric, 43).value
            Shopping = ws.cell(row_numeric, 44).value
            OtherPublicPlaces = ws.cell(row_numeric, 45).value

            row_numeric += 1

            #doing valid file calculation first
            if Sleeping == None: #if they were awake
                if Problem == None: #and there was no problem
                    Valid_counter += 1
                    VALID_FILE = True
                if Problem != None and len(str(Problem)) > 0 and E_text_cell != None and E_text_cell != " ": #or if there was a problem but the recorder still captured audio
                    if E_text_cell != None:
                        if len(E_text_cell) == 1:
                            if E_text_cell == " ":
                                print("SOMEONE PUT A SPACE :( (E), line ", i)
                            else:
                                print("It's something else (E)",i, len(E_text_cell),E_text_cell)
                    if NE_text_cell != None:
                        if len(NE_text_cell) == 1:
                            if NE_text_cell == " ":
                                print("SOMEONE PUT A SPACE :( (NE), line ", i)
                            else:
                                print("It's something else (NE)",i, len(NE_text_cell),NE_text_cell)
                    Valid_counter += 1
                    VALID_FILE = True
            if Sleeping != None and type(Sleeping) == int:
                VALID_FILE = False
            if Problem in [1,2,3,4] and E_text_cell == None:
                VALID_FILE = False
            if VALID_FILE == True:
                self.valid_counter_two += 1

            if E_text_cell != None: #if there's words in the cell:
                E_text_exists = True
                E_clean_text = clean_up(E_text_cell) #cleans up the raw text, returns a list where each element is a word
                if len(E_clean_text) == 0:
                    E_clean_text_exists = False
                else:
                    E_clean_text_exists = True
                self.E_words_per_chunk_list.append(len(E_clean_text)) #keeps track of how many words each person says per sample
                if len(E_clean_text) > 0: #some are empty after cleaning, don't need those, will append 0 to the words per chunk list to track that
                    self.lines_E_text[row_numeric] = E_clean_text
                    E_lines_list.append(E_clean_text) #append the text list to the lines list (only keeps track of the text)
                    self.E_words_per_chunk_with_speech.append(len(E_clean_text)) #length of the list is the number of words
                    self.Both_words_per_chunk_with_speech.append(len(E_clean_text)) #for keeping track of the length of E & NE
                    for word in E_clean_text: #for each word in the list:
                        self.E_one_list.append(word) #add it to the list that is just one big list of each word instead of a list of lists
            else:
                self.E_words_per_chunk_list.append(0) #if there were no words append that
                E_text_exists = False
                E_clean_text_exists = False

            #now do the same steps for the translation
            if NE_text_cell != None:
                NE_text_exists = True
                NE_clean_text = clean_up_NE(NE_text_cell) #clean up, return list
                self.NE_words_per_chunk_list.append(len(NE_clean_text)) #length of list = number of words (can be zero)
                if len(NE_clean_text) > 0: #if there are words after cleaning (meaning if they said words)
                    NE_clean_text_exists = True
                    self.lines_NE_text[row_numeric] = NE_clean_text
                    NE_lines_list.append(NE_clean_text) #count them
                    self.NE_words_per_chunk_with_speech.append(len(NE_clean_text)) #and add that to the list
                    self.Both_words_per_chunk_with_speech.append(len(NE_clean_text)) #also add it to the one that's counting both
                    for word in NE_clean_text:
                        self.NE_one_list.append(word) #add it to the list that is just one big list of each word instead of a list of lists
                else:
                    NE_clean_text_exists = False
            else:
                self.NE_words_per_chunk_list.append(0) #no text at all
                NE_text_exists = False
                NE_clean_text_exists = False

            self.Total_words_per_chunk_list.append(len(self.E_words_per_chunk_list)+len(self.NE_words_per_chunk_list)) #words in both english and 2nd language

            #Day of week general calcs
            if Day_of_week == 1: #if weekday
                WEEKDAY = True
                WEEKEND = False
                self.num_weekday_files += 1 #increment total number of files for the week
                if VALID_FILE == True:
                    self.num_valid_weekday_files += 1
            if Day_of_week == 2: #if weekend
                WEEKDAY = False
                WEEKEND = True
                self.num_weekend_files += 1
                if VALID_FILE == True:
                    self.num_valid_weekend_files += 1

            if WEEKDAY == True:
                if E_text_cell != None: #if there's text in that file
                    self.Overall_week_speech_files += 1 #if there was anything recorded, it's always in the raw transcription column, so this is how to count overall files with some sort of speech

                    #English files with speech, the number of tokens, and saving the text to count types later
                    if len(E_clean_text) > 0: #if the cleaned text isn't just codes/bkg noises
                        self.Enum_week_files_with_speech += 1 #increment counter of number of files that have actual speech
                        self.Eweek_files += E_clean_text #add the cleaned text to a list keeping track of the week text specifically
                        self.Enum_week_tokens += len(E_clean_text) #add the number of words to the total keeping track of tokens for the week specifically

                    if E_clean_text_exists == False and NE_text_cell != None:
                        if len(NE_clean_text) >0:
                            self.Only_NE_Week_Files_With_Speech += 1

                #Number of NE
                if NE_text_cell != None:
                    if len(NE_clean_text) > 0:
                        self.NEnum_week_files_with_speech += 1
                        self.NEweek_files += NE_clean_text #keeping track of the tokens so you can go and make a types dictionary later
                        self.NEnum_week_tokens += len(NE_clean_text)

                if NE_text_cell != None and E_text_cell != None: #if they spoke in both langs
                    if len(NE_clean_text) >0 and len(E_clean_text) >0: #and it wasn't just backg noise
                        self.Num_week_files_with_speech_both += 1 #they codeswitched

                if E_text_cell != None and NE_text_exists == False: #if there's text in the transcription column but nothign translated
                    if len(E_clean_text) >0:
                        self.Only_E_Week_Files_With_Speech += 1 #then they only spoke english

                if E_text_cell != None and NE_text_exists == True:
                    if len(E_clean_text) > 0 and len(NE_clean_text) == 0: #or if the above conditions happened but whatever was in the NE text was garbage
                        self.Only_E_Week_Files_With_Speech += 1

            if WEEKEND == True: #then do the same thing but for the weekend files
                if E_text_cell != None:
                    self.Overall_weekend_speech_files += 1 #if there was anything recorded, it's always in the raw transcription column, so this is how to count overall files with some sort of speech

                    if len(E_clean_text) > 0:
                        self.Enum_weekend_files_with_speech+=1 #increment counter of number of English files that have speech
                        self.Eweekend_files += E_clean_text #keeping track of the tokens so you can go and make a types dictionary later
                        self.Enum_weekend_tokens += len(E_clean_text)

                        #if self.file_name == "Ea"
                    if E_clean_text_exists == False and NE_text_cell != None:
                        if len(NE_clean_text) >0:
                            self.Only_NE_Weekend_Files_With_Speech += 1

                if NE_text_cell != None:
                    if len(NE_clean_text) > 0:
                        self.NEnum_weekend_files_with_speech += 1
                        self.NEweekend_files += NE_clean_text
                        self.NEnum_weekend_tokens += len(NE_clean_text)

                if E_text_cell != None and NE_text_cell != None:
                    if len(NE_clean_text) > 0 and len(E_clean_text) > 0:
                        self.Num_weekend_files_with_speech_both += 1

                if E_text_cell != None and NE_text_exists == False: #if there's stuff in the english transcription column but not in the NE column
                    if len(E_clean_text) >0: #and if it's not garbage
                        self.Only_E_Weekend_Files_With_Speech += 1

                if E_text_cell != None and NE_text_exists == True:
                    if len(E_clean_text) > 0 and len(NE_clean_text) == 0:
                        self.Only_E_Weekend_Files_With_Speech += 1

            #calculating locaton totals added 3/14/2022
            if E_text_cell != None or NE_text_cell != None: #if there was something in the transcript or translation columns
                if E_clean_text_exists == True or NE_clean_text_exists == True: #first check if it made it to clean text
                    if len(E_clean_text) > 0 or len(NE_clean_text) > 0: #should be analogous to cases where Talk column would be 1 - ie there are words left over in either the English, Nonenglish, or both columns after cleaning
                        if Outdoors != None:
                            self.num_Outdoor_files += 1
                        if AptDorm != None:
                            self.num_AptDorm_files += 1
                        if Classroom != None:
                            self.num_Classroom_files += 1
                        if InTransitVehicle != None:
                            self.num_InTransitVehicle_files += 1
                        if InTransitOther != None:
                            self.num_IntransitOther_files += 1
                        if BarCoffeeShopRestaurant != None:
                            self.num_BarCoffeeShopRestaurant_files += 1
                        if Shopping != None:
                            self.num_Shopping_files += 1
                        if OtherPublicPlaces != None:
                            self.num_OtherPublicPlaces_files += 1
                        if Outdoors == None and AptDorm == None and Classroom == None and InTransitVehicle == None and InTransitOther == None and BarCoffeeShopRestaurant == None and Shopping == None and OtherPublicPlaces == None:
                            self.num_NoLocation_files += 1

                        if MF != None:
                            if MF == 1:
                                self.num_ToM_files += 1
                            if MF == 2:
                                self.num_ToF_files += 1
                            if MF == 3:
                                self.num_ToMF_files += 1
                        if Self != None:
                            self.num_ToSelf_files += 1
                        if KnownPerson != None:
                            self.num_ToKnown_files += 1
                        if Stranger != None:
                            self.num_ToStranger_files += 1
                        if Child != None:
                            self.num_ToChild_files += 1
                        if Pet != None:
                            self.num_ToPet_files += 1
                        if MF == None and Self == None and KnownPerson == None and Stranger == None and Child == None and Pet == None:
                            self.num_NoKnownConvoPartner_files += 1
            else:
                print("Neither E nor NE clean text")

        #making week & weekend dicts for E and NE speech (function is just the basic type dictionary thing but I got tired of writing it out every time)
        self.Eweek_types_dict = make_dict_from_tokens(self.Eweek_files)
        self.Eweekend_types_dict = make_dict_from_tokens(self.Eweekend_files)
        self.NEweek_types_dict = make_dict_from_tokens(self.NEweek_files)
        self.NEweekend_types_dict = make_dict_from_tokens(self.NEweekend_files)

        #taking the number of keys in the dict to get the total number of types
        self.Enum_week_types = len(list(self.Eweek_types_dict.keys()))
        self.NEnum_week_types = len(list(self.NEweek_types_dict.keys()))
        self.Enum_weekend_types = len(list(self.Eweekend_types_dict.keys()))
        self.NEnum_weekend_types = len(list(self.NEweekend_types_dict.keys()))

        self.num_valid_files = Valid_counter

        #after you're done going through each cell & tracking if there's text or not, cleaning the text, and all that stuff, do averages for the participant
        self.E_avg_num_words_per_sample = np.mean(self.E_words_per_chunk_with_speech)
        self.NE_avg_num_words_per_sample = np.mean(self.NE_words_per_chunk_with_speech)
        self.Overall_avg_num_words_per_sample = np.mean(self.Both_words_per_chunk_with_speech) #I think this is the right way to do that?

        self.E_num_lines_with_text = len(E_lines_list)
        self.NE_num_lines_with_text = len(NE_lines_list)

        self.E_raw_text_list = E_lines_list #raw text as list of list (each list is one of the excel cells with all the individual words as the elements of the list)
        self.NE_raw_text_list = NE_lines_list

        self.E_raw_text_string = make_list_a_string(E_lines_list) #raw text as one big string
        self.NE_raw_text_string = make_list_a_string(NE_lines_list)

        for item in self.lines_E_text.items():
            if item[0] not in self.lines_total:
                self.lines_total[item[0]] = 1
            else:
                print("Something's wrong")
                pass #shouldn't even get here b/c empty dict
        for item in self.lines_NE_text.items():
            if item[0] not in self.lines_total:
                self.lines_total[item[0]] = 1
            else:
                self.lines_total[item[0]] += 1 #should never get above 2

        self.num_files_any_speech_E_NE = len(list(self.lines_total.keys()))

        return self.E_raw_text_list, self.E_raw_text_string, self.E_num_lines_with_text, self.E_avg_num_words_per_sample, self.E_one_list, self.NE_raw_text_list, self.NE_raw_text_string, self.NE_num_lines_with_text, self.NE_avg_num_words_per_sample, self.NE_one_list, self.num_valid_files, self.num_files_any_speech_E_NE

    def get_types_and_tokens(self, raw_text_list):
        self.num_types = 0
        self.num_tokens = 0
        self.types_dict = {}
        self.tokens_dict = {}
        for words_list in raw_text_list: #for each list of words (the excel cell gets turned into a list)
            for word in words_list: #for each word in that list
                if word not in self.types_dict: #if the word is already in the types dictionary
                    self.types_dict[word] = 1 #add it if not already there
                    self.num_types += 1 #and also increment (do nothing if it is there because that's tokens not types)
                if word in self.tokens_dict:
                    self.tokens_dict[word] += 1
                    self.num_tokens += 1
                else:
                    self.tokens_dict[word] = 1
                    self.num_tokens += 1

        return self.num_types, self.types_dict, self.num_tokens, self.tokens_dict

    def get_types_and_tokens_NE(self, NE_raw_text_list):
        types = 0
        tokens = 0
        types_dict = {}
        tokens_dict = {}
        for words_list in NE_raw_text_list: #for each list of words (the excel cell gets turned into a list)
            for word in words_list: #for each word in that list
                if word not in types_dict: #if the word is already in the types dictionary
                    types_dict[word] = 1 #add it if not already there
                    types += 1 #and also increment (do nothing if it is there because that's tokens not types)
                if word in tokens_dict:
                    tokens_dict[word] += 1
                    tokens += 1
                else:
                    tokens_dict[word] = 1
                    tokens += 1
        self.NE_num_types = types
        self.NE_types_dict = types_dict
        self.NE_num_tokens = tokens
        self.NE_tokens_dict = tokens_dict

        return self.NE_num_types, self.NE_types_dict, self.NE_num_tokens, self.NE_tokens_dict

    def types_in_500(self, raw_text_list):
        types = {}
        for token in raw_text_list[0:500]: #for each word in the first 500 words of that participant's list of words
            if token not in types: #if that's not already in the types dictionary
                types[token] = 1 #it is now
        num_types_in_five_hundred = int(sum(types.values())) #number of unique words the particpant said in the first 500 words
        return types, num_types_in_five_hundred

    #repeat at 1000 tokens, 2000 tokens, and the entire text (if that's larger)
    def types_in_1k(self, raw_text_list):
        types = {}
        for token in raw_text_list[0:1000]:
            if token not in types:
                types[token] = 1
        num_types_in_1000 = int(sum(types.values()))
        return types, num_types_in_1000

    def types_in_2k(self, raw_text_list):
        types = {}
        for token in raw_text_list[0:2000]:
            if token not in types:
                types[token] = 1
        num_types_in_2000 = int(sum(types.values()))
        return types, num_types_in_2000

    def types_in_however_many(self, raw_text_list):
        types = {}
        for token in raw_text_list:
            if token not in types:
                types[token] = 1
        num_types_in_all = int(sum(types.values()))
        return types, num_types_in_all

def main():
    all_english_text = ""
    all_non_english_text = ""

    files_list = get_file_names()
    files_class_list = create_TextFile_classes(files_list) #creating class instances for the text files
    num_participants = len(files_class_list)
    c = 1
    for file_class in files_class_list: #for each class instance
        name = file_class.file_name #file name so it knows which file to
        participant = str(round(int((name[4:-5])))) #cutting off file formatting so it's just the participant number

        #get the text & translation text
        E_raw_text_list, E_raw_text_string, E_num_lines_with_text, E_avg_num_words_per_sample, E_one_list, NE_raw_text_list, NE_raw_text_string, NE_num_lines_with_text, NE_avg_num_words_per_sample, NE_one_list, numvalidfiles, num_files_any_speech_E_NE = file_class.read_file(name)

        print(participant)#, num_files_any_speech_E_NE)
        print(((c/num_participants)*100), "% done")
        #Getting types and tokens
        E_num_types, E_types_dict, E_num_tokens, E_tokens_dict = file_class.get_types_and_tokens(E_raw_text_list)
        NE_num_types, NE_types_dict, NE_num_tokens, NE_tokens_dict = file_class.get_types_and_tokens_NE(NE_raw_text_list)

        #text from list of words to one long string
        E_all_text = ' '.join(E_one_list) #takes the list of individual words and makes them all a string (easier for tagging) - it's just for that one participant (class instance)
        NE_all_text = ' '.join(NE_one_list)
        all_english_text += E_all_text
        all_non_english_text += NE_all_text

        #types and tokens for E, NE, & both together (Dictionaries are at gloabal level, this is just updating them with a new kvp for each participant)
        E_Tokens_Dict[participant] = int(E_num_tokens) #dictionary of the number of types and tokens each participant said (not actual text, just the numberr)
        E_Types_Dict[participant] = int(E_num_types)
        NE_Tokens_Dict[participant] = int(NE_num_tokens)
        NE_Types_Dict[participant] = int(NE_num_types)
        E_NE_Tokens_Dict[participant] = int(NE_num_tokens+E_num_tokens)
        E_NE_Types_Dict[participant] = int(NE_num_types+E_num_types) #Is this the best way to do that?

        #Counting total number of rows in that participant's spreadsheet... not terribly interesting but could be a way to measure wearing on weekdays vs weekends?
        Total_Week_Files_Dict[participant] = int(file_class.num_weekday_files)
        Total_Weekend_Files_Dict[participant] = int(file_class.num_weekend_files)

        #Counting the number of total valid files on weekday vs weekend.
        Total_Valid_Week_Files_Dict[participant] = int(file_class.num_valid_weekday_files)
        Total_Valid_Weekend_Files_Dict[participant] = int(file_class.num_valid_weekend_files)

        #Number of files with speech in English
        Total_E_Week_Files_With_Speech[participant] = int(file_class.Enum_week_files_with_speech)
        Total_E_Weekend_Files_With_Speech[participant] = int(file_class.Enum_weekend_files_with_speech)

        #Number of files with speech not in English
        Total_NE_Week_Files_With_Speech[participant] = int(file_class.NEnum_week_files_with_speech)
        Total_NE_Weekend_Files_With_Speech[participant] = int(file_class.NEnum_weekend_files_with_speech)

        #Number of speech files only English or only Non-English
        Only_E_Week_Files_With_Speech[participant] = int(file_class.Only_E_Week_Files_With_Speech)
        Only_E_Weekend_Files_With_Speech[participant] = int(file_class.Only_E_Weekend_Files_With_Speech)
        Only_NE_Week_Files_With_Speech[participant] = int(file_class.Only_NE_Week_Files_With_Speech)
        Only_NE_Weekend_Files_With_Speech[participant] = int(file_class.Only_NE_Weekend_Files_With_Speech)

        #Speech files where they codeswitched
        Weekday_Speech_Files_Both[participant] = int(file_class.Num_week_files_with_speech_both)
        Weekend_Speech_Files_Both[participant] = int(file_class.Num_weekend_files_with_speech_both)

        #Counting the number of files that have something in the transcription column
        Overall_Total_Week_Speech_Files[participant] = int(file_class.Overall_week_speech_files)
        Overall_Total_Weekend_Speech_Files[participant] = int(file_class.Overall_weekend_speech_files)

        #by weekday vs weekend, E & NE separated
        E_Weekday_Tokens_Dict[participant] = int(file_class.Enum_week_tokens) #dictionary of the number of types and tokens each participant said (not actual text, just the numberr)
        E_Weekday_Types_Dict[participant] = int(file_class.Enum_week_types)
        NE_Weekday_Tokens_Dict[participant] = int(file_class.NEnum_week_tokens)
        NE_Weekday_Types_Dict[participant] = int(file_class.NEnum_week_types)
        E_Weekend_Tokens_Dict[participant] = int(file_class.Enum_weekend_tokens) #dictionary of the number of types and tokens each participant said (not actual text, just the numberr)
        E_Weekend_Types_Dict[participant] = int(file_class.Enum_weekend_types)
        NE_Weekend_Tokens_Dict[participant] = int(file_class.NEnum_weekend_tokens)
        NE_Weekend_Types_Dict[participant] = int(file_class.NEnum_weekend_types)

        #Totals across E & NE for each participant for weekend & weekday
        E_NE_Weekday_Tokens_Dict[participant] = int(file_class.NEnum_week_tokens+file_class.Enum_week_tokens)
        E_NE_Weekday_Types_Dict[participant] = int(file_class.NEnum_week_types+file_class.Enum_week_types)
        E_NE_Weekend_Tokens_Dict[participant] = int(file_class.NEnum_weekend_tokens+file_class.Enum_weekend_tokens)
        E_NE_Weekend_Types_Dict[participant] = int(file_class.NEnum_weekend_types+file_class.Enum_weekend_types)

        #print({k: E_types_dict[k] for k in sorted(E_types_dict)})
        E_NE_NoDouble_Speech_Dict[participant] = int(num_files_any_speech_E_NE)

        E_Num_Files_With_Speech_Dict[participant] = int(E_num_lines_with_text)
        NE_Num_Files_With_Speech_Dict[participant] = int(NE_num_lines_with_text)

        E_Avg_Num_Words_Per_File_Dict[participant] = float(E_avg_num_words_per_sample)
        NE_Avg_Num_Words_Per_File_Dict[participant] = float(NE_avg_num_words_per_sample)

        #I switched methods of counting valid files partway through so this is just to make sure the old method counts the same as the new one
        Valid[participant] = int(numvalidfiles) #number of valid files per participant
        Valid_check[participant] = int(file_class.valid_counter_two)

        #locaton information added 3/14
        AptDorm_Dict[participant] = int(file_class.num_AptDorm_files)
        OutdoorFiles_Dict[participant] = int(file_class.num_Outdoor_files)
        Classroom_Dict[participant] = int(file_class.num_Classroom_files)
        InTransitVehicle_Dict[participant] = int(file_class.num_InTransitVehicle_files)
        InTransitOther_Dict[participant] = int(file_class.num_IntransitOther_files)
        BarCoffeeShopRestaurant_Dict[participant] = int(file_class.num_BarCoffeeShopRestaurant_files)
        Shopping_Dict[participant] = int(file_class.num_Shopping_files)
        OtherPublicPlaces_Dict[participant] = int(file_class.num_OtherPublicPlaces_files)
        NoLocation_Dict[participant] = int(file_class.num_NoLocation_files)

        #Conversation Partner information added 3/14/2022
        ToM_Dict[participant] = int(file_class.num_ToM_files)
        ToF_Dict[participant] = int(file_class.num_ToF_files)
        ToMF_Dict[participant] = int(file_class.num_ToMF_files)
        ToSelf_Dict[participant] = int(file_class.num_ToSelf_files)
        ToKnownPerson_Dict[participant] = int(file_class.num_ToKnown_files)
        ToStranger_Dict[participant] = int(file_class.num_ToStranger_files)
        ToChild_Dict[participant] = int(file_class.num_ToChild_files)
        ToPet_Dict[participant] = int(file_class.num_ToPet_files)
        NoInfoConversationPartner_Dict[participant] = int(file_class.num_NoKnownConvoPartner_files)

        #counting the number of types per 500, 1000, 2000 tokens in English and Non-English
        if E_num_tokens > 500:
            E_500_types, num_types_in_500 = file_class.types_in_500(E_one_list)
        else:
            num_types_in_500 = file_class.types_in_however_many(E_one_list)[1]
        if E_num_tokens > 1000:
            E_1000_types, num_types_in_1000 = file_class.types_in_1k(E_one_list)
        else:
            num_types_in_1000 = file_class.types_in_however_many(E_one_list)[1]
        if E_num_tokens > 2000:
            E_2000_types, num_types_in_2000 = file_class.types_in_2k(E_one_list)
        else:
            num_types_in_2000 = file_class.types_in_however_many(E_one_list)[1]
        E_types_in_all, E_num_types_all = file_class.types_in_however_many(E_one_list)
        E_1000_Dict[participant] = int(num_types_in_1000)
        E_2000_Dict[participant] = int(num_types_in_2000)
        E_500_Dict[participant] = int(num_types_in_500)

        if NE_num_tokens > 500:
            NE_500_types, NE_num_types_in_500 = file_class.types_in_500(NE_one_list)
        else:
            NE_500_types, NE_num_types_in_500 = file_class.types_in_however_many(NE_one_list)
        if NE_num_tokens > 1000:
            NE_1000_types, NE_num_types_in_1000 = file_class.types_in_1k(NE_one_list)
        else:
            NE_1000_types, NE_num_types_in_1000 = file_class.types_in_however_many(NE_one_list)
        if NE_num_tokens > 2000:
            NE_2000_types, NE_num_types_in_2000 = file_class.types_in_2k(NE_one_list)
        else:
            NE_2000_types, NE_num_types_in_2000 = file_class.types_in_however_many(NE_one_list)
        NE_types_all, NE_num_types_all = file_class.types_in_however_many(NE_one_list)
        NE_1000_Dict[participant] = int(NE_num_types_in_1000)
        NE_2000_Dict[participant] = int(NE_num_types_in_2000)
        NE_500_Dict[participant] = int(NE_num_types_in_500)

        c+=1
    #stats_csv_writer("word_stats_E_NE_weekend_sep17.csv", ["Participant","Total_English_Types","Total_Non-English_Types", "English_Tokens", "Non-English_Tokens", "Total_Tokens", "Englis_Types_Per_500_Tokens", "Non-English_Types_Per_500_Tokens", "English_Types_Per_1000_Tokens", "Non-English_Types_Per_1000_Tokens", "Englis_Types_Per_2000_Tokens", "Non-English_Types_Per_2000_Tokens", "Number_Of_English_Speech_Files", "Number_Of_Non-English_Speech_Files", "Average_Number_Of_Words_Per_English_Speech_Sample","Average_Number_Of_Words_Per_Non-English_Speech_Sample", "Number_Of_Valid_Speech_Files","Double_checking_number_of_valid_speech_files", "Number_Of_Files_With_Speech_Total_Not_Double-counted" ,"Total_Number_Of_Weekday_Files_Recorded","Number_of_Valid_Weekday_Files","Number_Of_Total_Weekend_Files_Recorded","Total_Valid_Weekend_Files","E_Tokens_Weekday", "E_Types_Weekday", "NE_Tokens_Weekday", "NE_Types_Weekday", "E_Tokens_Weekend", "E_Types_Weekend", "NE_Tokens_Weekend", "NE_Types_Weekend", "E_NE_Tokens_Weekday", "E_NE_Types_Weekday", "E_NE_Tokens_Weekend", "E_NE_Types_Weekend", "Weekday_English_Files_with_actual_speech", "Weekend_English_Files_with_actual_speech", "Weekday_actual_speech_files_with_only_English_speech", "Weekend_actual_speech_files_with_only_English_speech", "Number_of_weekday_files_with_any_non-english_speech", "Number_of_weekend_files_with_any_non-english_speech", "Number_of_weekday_files_with_exclusively_non-english_speech", "Number_of_weekend_files_with_exclusively_non-english_speech", "Number_of_weekday_files_with_both_english_and_non-english", "Number_of_weekend_files_with_both_english_and_non-english", "Overall_week_speech_files_either_lang","Overall_weekend_speech_files_either_lang"], E_Types_Dict, NE_Types_Dict, E_Tokens_Dict, NE_Tokens_Dict, E_NE_Tokens_Dict, E_500_Dict, NE_500_Dict, E_1000_Dict, NE_1000_Dict, E_2000_Dict, NE_2000_Dict, E_Num_Files_With_Speech_Dict, NE_Num_Files_With_Speech_Dict, E_Avg_Num_Words_Per_File_Dict, NE_Avg_Num_Words_Per_File_Dict, Valid, Valid_check, E_NE_NoDouble_Speech_Dict, Total_Week_Files_Dict, Total_Valid_Week_Files_Dict, Total_Weekend_Files_Dict, Total_Valid_Weekend_Files_Dict, E_Weekday_Tokens_Dict, E_Weekday_Types_Dict, NE_Weekday_Tokens_Dict, NE_Weekday_Types_Dict, E_Weekend_Tokens_Dict, E_Weekend_Types_Dict, NE_Weekend_Tokens_Dict, NE_Weekend_Types_Dict, E_NE_Weekday_Tokens_Dict, E_NE_Weekday_Types_Dict, E_NE_Weekend_Tokens_Dict, E_NE_Weekend_Types_Dict, Total_E_Week_Files_With_Speech, Total_E_Weekend_Files_With_Speech, Only_E_Week_Files_With_Speech, Only_E_Weekend_Files_With_Speech, Total_NE_Week_Files_With_Speech, Total_NE_Weekend_Files_With_Speech, Only_NE_Week_Files_With_Speech, Only_NE_Weekend_Files_With_Speech, Weekday_Speech_Files_Both, Weekend_Speech_Files_Both, Overall_Total_Week_Speech_Files, Overall_Total_Weekend_Speech_Files)
    stats_csv_writer("word_stats_E_NE_weekend_convopartners_location_3_27_2022.csv", LABEL_LIST, E_Types_Dict, NE_Types_Dict, E_Tokens_Dict, NE_Tokens_Dict, E_NE_Tokens_Dict, E_500_Dict, NE_500_Dict, E_1000_Dict, NE_1000_Dict, E_2000_Dict, NE_2000_Dict, E_Num_Files_With_Speech_Dict, NE_Num_Files_With_Speech_Dict, E_Avg_Num_Words_Per_File_Dict, NE_Avg_Num_Words_Per_File_Dict, Valid, Valid_check, E_NE_NoDouble_Speech_Dict, Total_Week_Files_Dict, Total_Valid_Week_Files_Dict, Total_Weekend_Files_Dict, Total_Valid_Weekend_Files_Dict, E_Weekday_Tokens_Dict, E_Weekday_Types_Dict, NE_Weekday_Tokens_Dict, NE_Weekday_Types_Dict, E_Weekend_Tokens_Dict, E_Weekend_Types_Dict, NE_Weekend_Tokens_Dict, NE_Weekend_Types_Dict, E_NE_Weekday_Tokens_Dict, E_NE_Weekday_Types_Dict, E_NE_Weekend_Tokens_Dict, E_NE_Weekend_Types_Dict, Total_E_Week_Files_With_Speech, Total_E_Weekend_Files_With_Speech, Only_E_Week_Files_With_Speech, Only_E_Weekend_Files_With_Speech, Total_NE_Week_Files_With_Speech, Total_NE_Weekend_Files_With_Speech, Only_NE_Week_Files_With_Speech, Only_NE_Weekend_Files_With_Speech, Weekday_Speech_Files_Both, Weekend_Speech_Files_Both, Overall_Total_Week_Speech_Files, Overall_Total_Weekend_Speech_Files, ToM_Dict, ToF_Dict, ToMF_Dict, ToSelf_Dict, ToKnownPerson_Dict, ToStranger_Dict, ToChild_Dict, ToPet_Dict, NoInfoConversationPartner_Dict, OutdoorFiles_Dict, AptDorm_Dict, Classroom_Dict, InTransitVehicle_Dict, InTransitOther_Dict, BarCoffeeShopRestaurant_Dict, Shopping_Dict, OtherPublicPlaces_Dict, NoLocation_Dict)

    #write_corpus_to_text(all_english_text, "EAR_english_corpus.txt")
    #write_corpus_to_text(all_non_english_text, "EAR_non_english_corpus.txt")
    #all_english_tokens_list = all_english_text.split() #one giant list of all the words
    #entire_E_corpus_as_dict_types, entire_E_corpus_as_dict_tokens = create_corpus_dict(all_english_tokens_list) #making a dictionary that just keeps track of the types in the corpus for later use
    #cd_counts_per_particpant = contexutal_diversity_participants(entire_E_corpus_as_dict_types, files_class_list) #contextual diversity count of number of participants who said each word
    #cd_counts_per_word_7 = contextual_diversity_window(entire_E_corpus_as_dict_types, all_english_tokens_list, window_size=WINDOW_SIZE) #contextual diversity counts for eachword

    #all particpants summary csvs - word counts and lexical diversity measures.
    #change_or_make_path("all_words_English")
    #all_words_and_cd_csv_writer("English_word_counts_and_contextual_diversity.csv",["word","count","number of people who said that word","number of words within a 7 word window"], entire_E_corpus_as_dict_tokens, cd_counts_per_particpant, cd_counts_per_word_7) #need the token one to get counts
    #words_grapher("English_word_counts_and_contextual_diversity.csv", sort_by_label="count", title="word counts", y_label = "counts")
    #words_grapher("English_word_counts_and_contextual_diversity.csv", sort_by_label="number of people who said that word", title="number of people who said that word", y_label ="number of people")
    #words_grapher("English_word_counts_and_contextual_diversity.csv", sort_by_label="number of words within a 7 word window", title="Unique words within a 7 word window", y_label="number of words")
    #os.chdir("..")

    #stacked_grapher(E_Types_Dict,NE_Types_Dict, "English and Non-English Types", "Particpant","Number of Types")
    #stacked_grapher(E_Tokens_Dict,NE_Tokens_Dict, "English and Non-English Tokens", "Participant", "Number of Tokens")
    #change_or_make_path("stats_both")
    #stacked_grapher(E_Weekday_Tokens_Dict, NE_Weekday_Tokens_Dict, title="English vs Non-English Weekday Speech", x_lab=" ", y_lab="Tokens")
    #stacked_grapher(E_Weekend_Tokens_Dict, NE_Weekend_Tokens_Dict, title="English vs Non-English Weekend Speech", x_lab=" ", y_lab="Tokens")
    #stacked_grapher_general(E_Weekday_Tokens_Dict, E_Weekend_Tokens_Dict, title="English Weekday vs Weekend Speech", x_lab=" ", y_lab="Tokens", dd1_name="Weekday Tokens", dd2_name="Weekend Tokens")
    #stacked_grapher_general(NE_Weekday_Tokens_Dict, NE_Weekend_Tokens_Dict, title="Non-English Weekday vs Weekend Speech", x_lab=" ", y_lab="Tokens", dd1_name="Weekday Tokens", dd2_name="Weekend Tokens")

    #participant_grapher("word_stats_E_NE_weekend_sep16.csv", sort_by_label="Diff_clean_week-weekend", title="Difference in Total Number of Files With (CLEAN) Speech on Weekdays - Weekends", y_label = "Difference", col = 64)
    #participant_grapher("word_stats_E_NE_weekend_sep16.csv", sort_by_label="Diff_tokens_Eweek-Eweekend", title="Difference in number of English Tokens on Weekdays - Weekends", y_label = "Difference", col = 65)
    #participant_grapher("word_stats_E_NE_weekend_sep16.csv", sort_by_label="Diff_tokens_NEweek-NEweekend", title="Difference in number of Non-English Tokens on Weekdays - Weekends", y_label = "Difference", col = 66)
    #participant_grapher("word_stats_E_NE_weekend_sep16.csv", sort_by_label="Diff_total_tokens_weekday-weekend", title="Difference in Total Number of Tokens on Weekdays - Weekends", y_label = "Difference", col = 67)
    #participant_grapher("word_stats_E_NE_weekend_sep16.csv", sort_by_label="Diff_files_with_any_sound_weekday-weekend", title="Difference in number of Files with any sound Weekdays - Weekends", y_label = "Difference", col = 68)
    #participant_grapher("word_stats_E_NE_weekend_sep16.csv", sort_by_label="Diff_valid_files_weekday-weekend", title="Difference in Total Number of Valid Files Weekdays - Weekends", y_label = "Difference", col = 69)
    #stacked_grapher_proportion(E_Weekday_Tokens_Dict, E_Weekend_Tokens_Dict, NE_Weekday_Tokens_Dict, NE_Weekend_Tokens_Dict, num_bars=4, title="English and Non-English Speech on Weekdays and Weekends", y_lab="tokens", figname="proportions.png")
    #proportions_grapher("word_stats_E_NE_weekend_sep16.csv", sum_label="add_props", t="proportion of speech in English and Non-English Languages on the weekend and weekdays", y_label= "proportion", col = [0,58,59,60,61])
main()
