#Project program script
import os
import re
import csv
import copy
from openpyxl import load_workbook #prior versions used xlrd, which at some point decided to only read xls files and no longer support xlsx. That is our data file format, hence the change


PUNCT = ["?","!" ,"@" ,"," ,"-" ,"(", ")" ,"[", "]", ";", ":", "?", ",",'']

#getting user input of the words they want to look at
def get_user_words():
    SPECIFIC_WORDS = []
    DONE = False
    print("Please enter each word you'd like to look at, one at a time. Please type DONE when you have entered all words")
    while DONE == False:
        words = input("Please type a word: ")
        if words == "DONE":
            DONE = True
            break
        SPECIFIC_WORDS.append(words)
        #print("Done status:", DONE)
    return(SPECIFIC_WORDS)

def make_list_of_dicts(list_of_words): #makes dictionaries for each of the words you want to track the line number and text
    dict_list = []
    for word in list_of_words:
        word = dict()
        dict_list.append(word)
    return dict_list

def get_file_names(): #names of all files that are English transcripts to be used later
    files_list = []
    current_wd = os.getcwd()
    os.chdir(current_wd+"/workbooks")
    files = os.listdir()
    for file in files:
        if file[0] != '.' and file[0] == 'E': #don't want the .DS store thing or any of the output csvs if you've already run this
            files_list.append(file)
    return files_list

def create_TextFile_classes(files_list): #creates class objects of each file & puts them in a list to be used for easy access later
    file_class_list = []
    for file in files_list:
        file_name = file
        file_class = TextFile(file_name)
        file_class_list.append(file_class)
    return file_class_list

def clean_up(text_file):
    '''Function to clean text input. I'm sorry it's a mess. Change lines as needed based on your analyses and coding scheme.
    Also yes I know there's a bunch of regular expressions that can be combined but for our purposes, L2 speech was coded as a 3 character string of a bunch of different consonants, so removing any combination of the letters in the tags led to segments of words being deleted :('''

    text = text_file.replace("\n"," ")
    text = text.replace('\\','') #could't make this work with regular expressions, got a bad escape error
    text = re.sub("[\[].*?[\]]", "", text) #removing everything in brackets, parenthese, or curly brackets
    text = re.sub("[\{].*?[\}]","", text)
    text = re.sub("[\(].*?[\)]","", text)
    text = re.sub("[\<].*?[\>]","", text)
    text = re.sub("[\[\{].*?[\}\]]","",text)
    text = re.sub("sss","",text)
    text = re.sub("sxx","",text)
    text = re.sub("tss","",text)
    text = re.sub("rrr","",text)
    text = re.sub("kkk","",text)
    text = re.sub("ppp","",text)
    text = re.sub("eee","",text)
    text = re.sub("cxc","",text)
    text = re.sub("ggg","",text)
    text = re.sub("mmm","",text)
    text = re.sub("Sxs","",text)
    text = re.sub("Sss","",text)
    text = re.sub("Ccc","",text)
    text = re.sub("xpp","",text)
    text = re.sub("xpx","",text)
    text = re.sub("xrr","",text)
    text = re.sub("xmm","",text)
    text = re.sub("jjj","",text)
    text = re.sub("jxj","",text)
    text = re.sub("fxf","",text)
    text = re.sub("mmm","",text)
    text = re.sub("mxm","",text)
    text = re.sub("bbb","",text)
    text = re.sub("vvv","",text)
    text = re.sub("ccc","",text)
    text = re.sub("xkx","",text)
    text = re.sub("kxk","",text)
    text = re.sub("xaa","",text)
    text = re.sub("xax","",text)
    text = re.sub("axa","",text)
    text = re.sub("rrr","",text)
    text = re.sub("txt","",text)
    text = re.sub("xtt","",text)
    text = re.sub("ttt","",text)
    text = re.sub("SSS","",text)
    text = re.sub("fff","",text)
    text = re.sub("aaa","",text)
    text = re.sub("eee","",text)
    text = re.sub("xjx","",text)
    text = re.sub("xxss","",text)
    text = re.sub("xss","",text)
    text = re.sub("mmx","",text)
    text = re.sub("xvx","",text)
    text = re.sub("xvv","",text)
    text = re.sub("xff","",text)
    text = re.sub("xsx","",text)
    text = re.sub("sxs","",text)
    text = re.sub("ssx","",text)
    text = re.sub("sss","",text)
    text = re.sub("Ttt","",text)
    text = re.sub("Aaa","",text)
    text = re.sub("Eee","",text)
    text = re.sub("bxb","",text)
    text = re.sub("6","six",text)
    text = re.sub("3","three",text)
    text = re.sub(" 2 "," two ",text)
    text = re.sub("4","four",text)
    text = re.sub("5","five",text)
    text = re.sub(" 1 "," one ",text)
    text = re.sub("21st","twenty-first",text)
    text = re.sub(" '", " ", text)
    text = re.sub("-"," ",text)
    text = re.sub(";","",text)
    text = re.sub("[\*]{3,}","", text) #removing all stars (change to {3,} if you want to keep the like*)
    text = re.sub("[\#X]","", text)
    text = re.sub("'","'",text) #??? what is that first character doing
    text = re.sub("\."," ", text)
    text = re.sub("。","",text)
    text = re.sub("…"," ",text)
    text = re.sub("' ", " ", text)
    text = re.sub(":","",text)
    text = re.sub("!","",text)
    text = re.sub("/","",text)
    text = re.sub("\n","", text) #removes newline characters because for some reason that was still an issue even with the first line
    text = re.sub("\?","",text)
    text = re.sub("é","e",text) #some people transcribed words like resume with an accent and excel did not like that
    text = re.sub(",","",text)
    text = re.sub("’","'", text)
    text = re.sub('"',"", text)
    #text = re.sub(r"[sfxvmjrkbctgpSFXVMJRKBCTGP]{3,}", " ", text)
    text = re.sub("[xx]{2,}"," ", text) #removing any number of sequential xs, as those were how conversation partner speech was noted
    text = re.sub("[aAxX]{3,}","",text)
    text = re.sub("[Aa]{3,}","",text)
    text = re.sub("[Ee]{3,}","",text)
    text = re.sub("'", " '", text) #hacky way of separating words with 's
    text = re.sub("\s+"," ", text) #removes multiple spaces and replaces with one space
    tokenized_text = text.split(" ")
    final = [word.lower().strip(" ") for word in tokenized_text if word not in PUNCT]
    return final

def change_or_make_path(path_addition):
    if os.path.exists(os.getcwd()+"/"+path_addition):
        os.chdir(os.getcwd()+"/"+path_addition)
    else:
        os.mkdir(path_addition)
        os.chdir(os.getcwd()+"/"+path_addition)

def nested_dictionary_csv_writer(words_list, nested_dictionary, file_name_list, label_list):
    for file_name in file_name_list: #write the first dict file

        with open(file_name,"w",newline="") as f:
            thewriter = csv.writer(f)
            thewriter.writerow(label_list)
            for item in nested_dictionary.items():
                participant = item[0]
                data = item[1]
                which_dictionary_in_data = data[file_name_list.index(file_name)] #get the dict to write
                for that_one in which_dictionary_in_data.items():
                    line_number = that_one[0]
                    text = that_one[1]
                    thewriter.writerow([participant,line_number,text])

def make_nested_dicts(list_of_participants, original_dicts_list):
    nested_dicts = dict()

    for participant in list_of_participants:
        participant = participant[4:7]
        nested_dicts[participant] = copy.deepcopy(original_dicts_list)

    print(type(nested_dicts))
    return nested_dicts

def get_cleaning_preference():
    DONE = False
    acceptable_answers = ["CLEANTEXT","RAWTEXT"]
    while DONE == False:
        clean_or_nah = str(input("Would you like your output file to consist of the raw transcripts, or the cleaned transcripts? \nIf you want punctuation and/or punctuation-based markers, or any transcript codes, you will want the raw text. \nIf you do not want any of those things and only want words, choose clean text. \nFor clean text type CLEANTEXT. For raw text type RAWTEXT. \n"))
        if clean_or_nah in acceptable_answers:
            DONE = True
        else:
            print("please type your answer exactly as it appears in the prompt")

    return clean_or_nah

class TextFile:
    def __init__(self, file_name):
        self.file_name = file_name
        self.raw_text_string = None
        self.raw_text_list = []
        self.words_per_chunk_list = []
        self.spacy_tagged_text = None
        self.num_types = 0
        self.num_tokens = 0
        self.adjs_dict_spacy = {}
        self.bigram_dict = {}

    def read_file(self, file_name, cleaning_preference, specific_words_list, nested_dicts): #reads the excel files - file name comes from files_list. Also has the dictionaries for the specific words you want. You'll need to change this if you want to look at more/different words
        file_name = file_name
        number = file_name[4:7]

        words_to_check = specific_words_list #the words your program is looking for occurrences of

        wb = load_workbook(filename = file_name, read_only=True) #load the excel file using openxyl
        ws = wb['Sheet1'] #go to sheet1 specifically
        cell_counter = 0
        start_cell = 1
        row_numeric = 2#needed due to differences in row/column indexing between xlrd and openpyxl

        current_dict_list = nested_dicts[number] #the list of dictionaries for the specific participant you're on

        for i in ws.iter_rows(min_row = 2, min_col = 3, max_col = 3):
            diff_counter = 0
            #row for file number - openpyxl includes some blank rows that xlrd did not, so making sure it stops at the end of the dataframe
            for cell in i:
                if cell.value != None:
                    if type(cell.value) == int:
                        if cell.value == start_cell:
                            start_cell+=1
                            #print("added 1, start cell now", start_cell)
                        else:
                            #print(start_cell)
                            diff = cell.value-start_cell
                            #print("the difference was: ", diff)
                            diff_counter += diff
                            total_diff = diff+1
                            start_cell += total_diff
                            #print("added the difference, start cell now", start_cell)
                    else:
                        print("File number value not an int", type(cell.value))
                else:
                    break
                cell_counter += 1
            start_cell -= diff_counter
        print("if using cell counter: ", cell_counter, "files for", file_name)
        print("using ", start_cell-1, "rows for file", file_name)
        for i in ws.iter_rows(min_row=2, max_row = start_cell): #start at second row to skip headers, go down column 5 for transcript
            cell_num = ws.cell(row= row_numeric, column = 3).value
            raw_text_E = ws.cell(row = row_numeric, column=5).value # column numbers are one off because this takes column numbers starting at 1, whereas xlrd used python indexing starting at 0
            raw_text_NE = ws.cell(row = row_numeric, column = 6).value
            if raw_text_E != None: #some are empty, don't need those
                if cleaning_preference == "CLEANTEXT":
                    clean_text_E = clean_up(raw_text_E) #cleans up the raw text, returns a list of strings of each word
                    search_in_text = clean_text_E
                if cleaning_preference == "RAWTEXT":
                    search_in_text = raw_text_E
                for word in words_to_check: #for each word you want to look at
                    re_method = re.findall(word, search_in_text)
                    print(re_method)
                    if word in search_in_text: #check for the word in the uncleaned sheet text in case there's any punctuation/formatting that needs to be retained
                        #print(word)
                        #get the corresponding dictionary for that word from your list
                        #pairs the index of the word in the list with the index of the dictionary in the dictionary list, so that you get all the 1st words in the 1st dictionary, 2nd words in the 2nd dictionary and so on
                        #take the corresponding dictionary for the word and then add a kvp with the line_number:line_text
                        #and then working in that dictionary, add the line number and the text associated with it for the participant
                        current_dict_list[words_to_check.index(word)][i] = search_in_text
        return current_dict_list

def main():
    words_list_to_use = get_user_words() #the words you want to get the lines of
    output_file_list = [] #for formatted names of your output CSVs
    dicts_list = make_list_of_dicts(words_list_to_use) #one dictionary for each of the words you're looking at
    files_list = get_file_names() #gets the participant excel file names to read
    cleaning_pref = get_cleaning_preference()
    nested_dict = make_nested_dicts(files_list, dicts_list) #one for each participant, and for each word you're looking at
    files_class_list = create_TextFile_classes(files_list) #creating class instances for the participant excel sheets

    for file_class in files_class_list: #for each class instance
        name = file_class.file_name #file name so it knows which file to read
        number = name[4:7] #number to pair it up with nested dictionary keys
        list_of_dicts = file_class.read_file(name, cleaning_pref, words_list_to_use, nested_dict) #get the list of dictionaries for each word for that participant
        nested_dict[number] = list_of_dicts #make them the value for the nested dictionary at the item of participant number

    for word in words_list_to_use:
        output_file_list.append(str(word)+".csv") #to make a file for each word you're looking at

    change_or_make_path("keywords")
    nested_dictionary_csv_writer(words_list_to_use, nested_dict, output_file_list, ["participant","file_number","text"])

main()
